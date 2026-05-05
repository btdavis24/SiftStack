"""Comparable sales analysis with Two-Bucket ARV methodology.

Generates appraiser-grade property valuations for real estate investment
analysis. Fetches comparable sales from the Zillow API, applies property-
specific adjustments, and produces a 7-tab Excel workbook.

Tennessee is a non-disclosure state — MLS/Zillow data is the primary
source, not public deed records.

Usage:
  python src/main.py comp --address "123 Main St, Knoxville, TN 37918"
  python src/main.py comp --address "123 Main St" --city Knoxville --zip 37918 --radius 0.5 --months 6
"""

import logging
import math
import random
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config

logger = logging.getLogger(__name__)

# ── API Configuration ─────────────────────────────────────────────────
API_BASE = "https://api.openwebninja.com/realtime-zillow-data"
PROPERTY_ENDPOINT = f"{API_BASE}/property-details-address"
PROPERTY_BY_ZPID_ENDPOINT = f"{API_BASE}/property-details"
COMPS_ENDPOINT = f"{API_BASE}/similar-sale-homes"
SEARCH_ENDPOINT = f"{API_BASE}/search"
REQUEST_DELAY_MIN = 1.0
REQUEST_DELAY_MAX = 2.0
REQUEST_TIMEOUT = 30
MAX_RETRIES = 2
MAX_SEARCH_PAGES = 12  # /search fallback: cap pagination (~3 weeks of sales per page)

# ── Comp selection defaults ───────────────────────────────────────────
DEFAULT_RADIUS_MILES = 0.5
MAX_RADIUS_MILES = 1.0
DEFAULT_MONTHS_BACK = 6
MAX_MONTHS_BACK = 12
MIN_COMPS = 3
TARGET_COMPS = 5
MAX_COMPS = 7

# ── Regional adjustment profiles ──────────────────────────────────────
# Per-unit adjustment amounts, calibrated per market. Selected by subject state.
# full_reno_premium is the $ delta between an as-is comp and a full-reno comp,
# used to convert comps to the subject's target-condition equivalent.
REGIONAL_ADJUSTMENTS = {
    "TN": {
        "sqft": 85.0,
        "bedroom": 10000.0,            # Recalibrated per skill: <$500k tier standard
        "bathroom": 10000.0,            # Recalibrated per skill
        "year_built": 500.0,
        "lot_sqft": 2.0,
        "lot_max": 15000.0,
        "garage": 12000.0,              # Recalibrated: skill says $10-15K standard
        "market_pct_mo": 0.003,         # 3.6% annual
        "full_reno_premium": 35000.0,   # East TN default — overridden by derivation when possible
        "bg_pct_of_ag": 0.40,
        "state_type": "disclosure",     # Per skill's state-list.md classification
        "region_label": "Knoxville / East Tennessee",
        "disclosure_note": "Tennessee is classified as disclosure per reference. Sold data via MLS/Zillow.",
    },
    "KY": {
        "sqft": 55.0,                   # ≈30% of Louisville PPSF ~$170
        "bedroom": 10000.0,             # Recalibrated per skill
        "bathroom": 10000.0,            # Recalibrated per skill
        "year_built": 500.0,
        "lot_sqft": 2.0,
        "lot_max": 15000.0,
        "garage": 12000.0,              # Recalibrated per skill
        "market_pct_mo": 0.002,         # ~2.4% annual, Louisville slower than Knoxville
        "full_reno_premium": 50000.0,   # Louisville default — overridden by derivation when possible
        "bg_pct_of_ag": 0.40,
        "state_type": "disclosure",     # KY is disclosure
        "region_label": "Louisville / Jefferson County, Kentucky",
        "disclosure_note": "Kentucky is a disclosure state. Sold prices are public record via MLS/county recorder.",
    },
}


def get_adj_profile(state: str) -> dict:
    """Return the regional adjustment profile for a state (defaults to TN)."""
    return REGIONAL_ADJUSTMENTS.get((state or "").upper(), REGIONAL_ADJUSTMENTS["TN"])


# ── Market phase detection ────────────────────────────────────────────
# Sentiment adjustment applied to arv_mid AFTER per-comp time-appreciation.
# Time adjustment captures "when this comp sold vs now"; sentiment captures
# "what buyers are doing right now" (urgency / hesitancy premium).
# DOM thresholds from skill's reference; multipliers kept conservative (±3-5%).
MARKET_PHASE_DOM_THRESHOLDS = [
    (15, "very_hot", 0.05),   # <15 avg DOM → very hot market, +5%
    (30, "hot", 0.03),        # 15-30 → hot, +3%
    (60, "balanced", 0.0),    # 30-60 → balanced, 0%
    (120, "cooling", -0.03),  # 60-120 → cooling, -3%
    (9999, "cool", -0.05),    # >120 → cool, -5%
]


def classify_market_phase(comps: list) -> tuple[str, int, float]:
    """Return (phase, avg_dom, sentiment_multiplier) from sold-comp DOM distribution.

    Uses top selected comps' days_on_market as the signal. Ignores comps with
    DOM=0 (off-market/FSBO sales that never posted to MLS).
    """
    doms = [c.days_on_market for c in comps if getattr(c, "days_on_market", 0) > 0]
    if not doms:
        return "balanced", 0, 0.0
    avg = sum(doms) / len(doms)
    for threshold, phase, mult in MARKET_PHASE_DOM_THRESHOLDS:
        if avg < threshold:
            return phase, int(round(avg)), mult
    return "balanced", int(round(avg)), 0.0


# ── Condition dimension ───────────────────────────────────────────────
# Renovation tier mapped to a 0-1 index. Condition adjustment = index delta ×
# profile["full_reno_premium"], so as-is→full moves a full premium, lighter
# tiers scale proportionally.
CONDITION_INDEX = {
    "as-is": 0.0, "as is": 0.0, "asis": 0.0, "fixer": 0.0, "needs-work": 0.0,
    "light": 0.35, "light-reno": 0.35, "light reno": 0.35,
    "between": 0.6, "between light and full": 0.6, "mid": 0.6,
    "near-full": 0.85, "near full": 0.85,
    "full": 1.0, "full-reno": 1.0, "full reno": 1.0, "renovated": 1.0,
}

# Keyword patterns used to auto-infer condition from Zillow description text.
# Order of precedence: as-is > full > light (as-is signals are most specific).
CONDITION_KEYWORDS = {
    "as-is": [
        "as-is", "as is", "fixer", "fixer upper", "investor special",
        "needs work", "needs tlc", "needs updating", "tlc needed",
        "handyman", "cash only", "bring your imagination",
        "diamond in the rough", "gut job",
    ],
    "full": [
        "fully renovated", "completely renovated", "completely updated",
        "totally remodeled", "gut renovated", "gut-renovated",
        "all new", "brand new", "new kitchen", "new baths",
        "turn-key", "turn key", "turnkey", "move-in ready",
    ],
    "light": [
        "updated", "new paint", "freshly painted", "recently painted",
        "refinished floors", "new flooring", "newer appliances",
    ],
}


def normalize_condition(label: str) -> float:
    """Return the 0-1 condition index for a label. Unknown → 0.5 (neutral)."""
    if not label:
        return 0.5
    return CONDITION_INDEX.get(label.lower().strip(), 0.5)


def detect_condition_from_description(desc: str) -> str:
    """Infer condition tier from Zillow's free-text description. Returns '' if no signal."""
    if not desc:
        return ""
    dl = desc.lower()
    # as-is tier checked first — its phrases are the least ambiguous
    for tier in ("as-is", "full", "light"):
        for kw in CONDITION_KEYWORDS[tier]:
            if kw in dl:
                return tier
    return ""


def load_condition_overrides(path: str) -> dict:
    """Load a zpid/address → override CSV.

    Recognized columns: zpid, address, condition, ag_sqft, bg_sqft, notes.
    Any row with at least one of (condition, ag_sqft, bg_sqft) is kept and
    indexed by both zpid (if present) and lowercased address. Returned dict
    maps `zpid:<id>` or `addr:<addr>` → {condition, ag_sqft, bg_sqft}.
    """
    import csv
    import os
    overrides = {}
    if not path or not os.path.exists(path):
        return overrides
    with open(path, encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            cond = (row.get("condition") or "").strip()
            ag = row.get("ag_sqft") or ""
            bg = row.get("bg_sqft") or ""
            try:
                ag_val = int(ag) if str(ag).strip() else 0
                bg_val = int(bg) if str(bg).strip() else 0
            except ValueError:
                ag_val, bg_val = 0, 0
            if not cond and not ag_val and not bg_val:
                continue
            entry = {"condition": cond, "ag_sqft": ag_val, "bg_sqft": bg_val}
            zpid = (row.get("zpid") or "").strip()
            addr = (row.get("address") or "").strip().lower()
            if zpid:
                overrides[f"zpid:{zpid}"] = entry
            if addr:
                overrides[f"addr:{addr}"] = entry
    return overrides

# ── Data structures ───────────────────────────────────────────────────


@dataclass
class SubjectProperty:
    """The property being analyzed."""
    address: str = ""
    city: str = ""
    state: str = "TN"
    zip_code: str = ""
    latitude: float = 0.0
    longitude: float = 0.0
    sqft: int = 0                 # Total finished (usually matches Zillow livingArea)
    ag_sqft: int = 0              # Above-grade finished
    bg_finished_sqft: int = 0     # Below-grade finished (basement, walk-out, bi-level lower)
    bedrooms: int = 0
    bathrooms: float = 0.0
    year_built: int = 0
    lot_sqft: int = 0
    property_type: str = ""
    zestimate: float = 0.0
    mls_status: str = ""
    last_sold_date: str = ""
    last_sold_price: float = 0.0
    garage_spaces: int = 0
    description: str = ""
    target_condition: str = "full"  # Condition the subject will be valued at


@dataclass
class CompProperty:
    """A comparable property with sale data."""
    zpid: str = ""
    address: str = ""
    city: str = ""
    state: str = "TN"
    zip_code: str = ""
    latitude: float = 0.0
    longitude: float = 0.0
    distance_miles: float = 0.0
    sqft: int = 0                 # Total finished (usually matches Zillow livingArea)
    ag_sqft: int = 0              # Above-grade finished (falls back to sqft if not split)
    bg_finished_sqft: int = 0     # Below-grade finished
    bedrooms: int = 0
    bathrooms: float = 0.0
    year_built: int = 0
    lot_sqft: int = 0
    property_type: str = ""
    sold_price: float = 0.0
    sold_date: str = ""
    days_on_market: int = 0
    garage_spaces: int = 0
    description: str = ""  # From /property-details enrichment; used for condition auto-detect
    condition: str = ""    # "as-is", "light", "full", etc. — override > auto-detect > ""
    condition_source: str = ""  # "override", "auto", or "" (unknown)
    # Calculated fields
    similarity_score: float = 0.0
    adjusted_price: float = 0.0
    ppsf: float = 0.0
    bucket: str = ""  # "A" (non-disclosure baseline) or "B" (disclosure/adjusted)
    adjustments: dict = field(default_factory=dict)


@dataclass
class ARVResult:
    """Final ARV calculation result."""
    arv_low: float = 0.0
    arv_mid: float = 0.0
    arv_high: float = 0.0
    arv_mid_pre_sentiment: float = 0.0  # Before market-sentiment multiplier applied
    confidence: str = ""  # "high", "medium", "low"
    confidence_reason: str = ""
    ppsf_avg: float = 0.0
    ppsf_range: tuple = (0.0, 0.0)
    comp_count: int = 0
    bucket_a_count: int = 0
    bucket_b_count: int = 0
    avg_adjustment: float = 0.0
    spread_pct: float = 0.0
    # Market sentiment layer (separate from per-comp time-based appreciation)
    market_phase: str = "balanced"    # "very_hot" / "hot" / "balanced" / "cooling" / "cool"
    market_dom_avg: int = 0
    sentiment_adj_pct: float = 0.0    # e.g., 0.03 for +3%, -0.05 for -5%
    reno_premium_used: float = 0.0    # Actual $ used (derived or fallback)
    reno_premium_source: str = ""     # "derived" or "fallback"


# ── Distance calculation ──────────────────────────────────────────────

def _safe_int(val) -> int:
    """Parse Zillow numeric fields that may arrive as int, float, or comma-formatted strings."""
    if val is None or val == "":
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(float(str(val).replace(",", "").strip()))
    except (ValueError, TypeError):
        return 0


def _haversine_miles(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate distance between two lat/lon points in miles."""
    R = 3958.8  # Earth radius in miles
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))


# ── API calls ─────────────────────────────────────────────────────────

def _api_get(endpoint: str, params: dict, api_key: str) -> dict | None:
    """Make an authenticated GET request to OpenWeb Ninja API."""
    headers = {"x-api-key": api_key}
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(endpoint, headers=headers, params=params, timeout=REQUEST_TIMEOUT)
            if resp.status_code == 404:
                return None
            if resp.status_code == 429:
                logger.warning("Rate limit hit — waiting 10s (attempt %d)", attempt)
                time.sleep(10)
                continue
            resp.raise_for_status()
            body = resp.json()
            if body.get("status") == "OK" and body.get("data"):
                return body["data"]
            return body if isinstance(body, list) else None
        except requests.Timeout:
            logger.warning("Timeout (attempt %d/%d)", attempt, MAX_RETRIES)
        except requests.RequestException as e:
            logger.warning("API error: %s (attempt %d/%d)", e, attempt, MAX_RETRIES)
    return None


def fetch_subject_property(address: str, city: str = "", state: str = "TN",
                           zip_code: str = "", api_key: str = "",
                           overrides: dict | None = None,
                           target_condition: str = "full") -> SubjectProperty | None:
    """Fetch full property details for the subject property.

    `overrides` lets the caller correct Zillow's reso facts (sometimes stale):
    keys: beds, baths, sqft, year_built, lot_sqft, garage, property_type.
    """
    api_key = api_key or config.OPENWEBNINJA_API_KEY
    if not api_key:
        logger.error("No OpenWeb Ninja API key configured")
        return None

    parts = [p for p in [address, city, state, zip_code] if p]
    full_address = " ".join(parts)

    data = _api_get(PROPERTY_ENDPOINT, {"address": full_address}, api_key)
    if not data:
        logger.warning("No property data found for '%s'", full_address)
        return None

    # Parse price history for last sold
    last_sold_date, last_sold_price = "", 0.0
    for entry in (data.get("priceHistory") or []):
        event = (entry.get("event") or "").lower()
        if event in ("sold", "listed (sold)"):
            last_sold_date = str(entry.get("date", ""))[:10]
            last_sold_price = float(entry.get("price") or 0)
            break

    # Parse lot size
    lot_sqft = 0
    lot_val = data.get("lotAreaValue")
    lot_units = (data.get("lotAreaUnits") or data.get("lotAreaUnit") or "").lower()
    if lot_val:
        lot_sqft = int(float(lot_val) * 43560) if "acre" in lot_units else int(float(lot_val))

    ov = overrides or {}
    reso = data.get("resoFacts") or {}
    auto_ag = _safe_int(reso.get("aboveGradeFinishedArea"))
    auto_bg = _safe_int(reso.get("belowGradeFinishedArea"))
    total_sqft = _safe_int(ov.get("sqft") or data.get("livingArea"))

    # AG: override wins, else resoFacts, else total (no split available)
    ag_sqft = int(ov.get("ag_sqft") or auto_ag or total_sqft)
    bg_sqft = int(ov.get("bg_sqft") if ov.get("bg_sqft") is not None else auto_bg)

    subject = SubjectProperty(
        address=data.get("streetAddress") or address,
        city=data.get("city") or city,
        state=data.get("state") or state,
        zip_code=str(data.get("zipcode") or zip_code),
        latitude=float(data.get("latitude") or 0),
        longitude=float(data.get("longitude") or 0),
        sqft=total_sqft,
        ag_sqft=ag_sqft,
        bg_finished_sqft=bg_sqft,
        bedrooms=int(ov.get("beds") or data.get("bedrooms") or 0),
        bathrooms=float(ov.get("baths") or data.get("bathrooms") or 0),
        year_built=int(ov.get("year_built") or data.get("yearBuilt") or 0),
        lot_sqft=int(ov.get("lot_sqft") or lot_sqft),
        property_type=ov.get("property_type") or data.get("homeType") or "",
        zestimate=float(data.get("zestimate") or 0),
        mls_status=data.get("homeStatus") or "",
        last_sold_date=last_sold_date,
        last_sold_price=last_sold_price,
        garage_spaces=int(ov.get("garage") if ov.get("garage") is not None else (data.get("garageSpaces") or 0)),
        description=data.get("description") or "",
        target_condition=target_condition or "full",
    )

    # Log any overrides that changed Zillow's values
    if overrides:
        changed = []
        for k in ("beds", "baths", "sqft", "ag_sqft", "bg_sqft", "year_built", "lot_sqft", "garage"):
            if ov.get(k) is not None:
                changed.append(f"{k}={ov[k]}")
        if changed:
            logger.info("Subject overrides applied: %s", ", ".join(changed))
    if auto_ag and auto_bg:
        logger.info("Subject AG/BG auto-detected from resoFacts: AG=%d, BG=%d", auto_ag, auto_bg)
    return subject


def _normalize_sold_date(item: dict) -> str:
    """Return sold date as YYYY-MM-DD. Handles string dates and epoch milliseconds."""
    raw = item.get("lastSoldDate") or item.get("dateSold")
    if not raw:
        return ""
    if isinstance(raw, (int, float)) or (isinstance(raw, str) and raw.isdigit()):
        # Epoch ms (Zillow /search returns this)
        ts = float(raw) / 1000.0
        try:
            return datetime.fromtimestamp(ts).strftime("%Y-%m-%d")
        except (OverflowError, OSError, ValueError):
            return ""
    return str(raw)[:10]


def _item_to_comp(item: dict, subject: SubjectProperty,
                  radius_miles: float, cutoff_date: str) -> CompProperty | None:
    """Convert one API result item to a CompProperty, applying date + distance filters.

    Returns None if the item should be excluded (no sold price, too old, too far).
    """
    if isinstance(item, str):
        return None

    sold_price = float(item.get("lastSoldPrice") or item.get("price")
                       or item.get("unformattedPrice") or 0)
    if not sold_price or sold_price < 10000:
        return None

    sold_date = _normalize_sold_date(item)
    if sold_date and sold_date < cutoff_date:
        return None

    lat = float(item.get("latitude") or 0)
    lon = float(item.get("longitude") or 0)
    dist = 0.0
    if subject.latitude and subject.longitude and lat and lon:
        dist = _haversine_miles(subject.latitude, subject.longitude, lat, lon)
        if dist > radius_miles:
            return None

    lot_sqft = 0
    lot_val = item.get("lotAreaValue") or item.get("lotSize")
    lot_units = (item.get("lotAreaUnits") or item.get("lotAreaUnit") or "").lower()
    if lot_val:
        try:
            lot_sqft = int(float(lot_val) * 43560) if "acre" in lot_units else int(float(lot_val))
        except (ValueError, TypeError):
            pass

    total_sqft = int(item.get("livingArea") or item.get("sqft") or 0)
    comp = CompProperty(
        zpid=str(item.get("zpid") or item.get("id") or ""),
        address=item.get("streetAddress") or item.get("address") or "",
        city=item.get("city") or "",
        state=item.get("state") or subject.state,
        zip_code=str(item.get("zipcode") or item.get("zip") or ""),
        latitude=lat,
        longitude=lon,
        distance_miles=round(dist, 2),
        sqft=total_sqft,
        ag_sqft=total_sqft,  # default AG = total; enrichment/override can split
        bg_finished_sqft=0,
        bedrooms=int(item.get("bedrooms") or 0),
        bathrooms=float(item.get("bathrooms") or 0),
        year_built=int(item.get("yearBuilt") or 0),
        lot_sqft=lot_sqft,
        property_type=item.get("homeType") or item.get("propertyType") or "",
        sold_price=sold_price,
        sold_date=sold_date,
        days_on_market=int(item.get("daysOnZillow") or 0),
        garage_spaces=int(item.get("garageSpaces") or 0),
    )
    comp.ppsf = round(comp.sold_price / comp.sqft, 2) if comp.sqft else 0.0
    return comp


def _fetch_sold_via_search(subject: SubjectProperty, radius_miles: float,
                           cutoff_date: str, api_key: str) -> list[CompProperty]:
    """Fallback path: paginate /search?home_status=RECENTLY_SOLD by ZIP code.

    Used when /similar-sale-homes returns nothing (common when the subject's Zillow
    record is sparse — e.g., homeStatus=OTHER or stale lastSoldDate).
    Stops paginating early once a page's results predate cutoff_date.
    /search results lack yearBuilt and garageSpaces; those adjustments are skipped.
    """
    location = subject.zip_code or f"{subject.city}, {subject.state}".strip(", ")
    if not location:
        logger.warning("No subject ZIP/city — cannot run sold search fallback")
        return []

    comps: list[CompProperty] = []
    for page in range(1, MAX_SEARCH_PAGES + 1):
        data = _api_get(SEARCH_ENDPOINT, {
            "location": location,
            "home_status": "RECENTLY_SOLD",
            "page": page,
        }, api_key)
        items = data if isinstance(data, list) else []
        if not items:
            break

        page_oldest = ""
        for item in items:
            if isinstance(item, dict):
                sd = _normalize_sold_date(item)
                if sd and (not page_oldest or sd < page_oldest):
                    page_oldest = sd
            comp = _item_to_comp(item, subject, radius_miles, cutoff_date)
            if comp:
                comps.append(comp)

        time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))

        # Stop paginating once the page's oldest sale is older than our cutoff
        if page_oldest and page_oldest < cutoff_date:
            break

    logger.info("Sold-search fallback: %d eligible comps from %d page(s) in '%s'",
                len(comps), page, location)
    return comps


def enrich_comps_with_details(comps: list[CompProperty], api_key: str = "") -> int:
    """Fill missing yearBuilt, garageSpaces, and description on comps via
    /property-details; auto-detect condition from the description text.

    The /search fallback returns sold listings without these fields, which means
    age/garage/condition adjustments can't be applied. One /property-details call
    per comp fills all three. Comps with a condition already set (from an override
    file) keep that value — we only auto-detect when condition is blank.
    Comps that already have both year_built AND a condition are skipped entirely.
    Returns the count of comps touched.
    """
    api_key = api_key or config.OPENWEBNINJA_API_KEY
    if not api_key:
        return 0

    enriched = 0
    for comp in comps:
        needs_yb = not comp.year_built
        needs_cond = not comp.condition
        # AG/BG split needed if comp still has bg=0 AND ag equals total (not split yet)
        needs_agbg = comp.bg_finished_sqft == 0 and comp.ag_sqft >= comp.sqft
        if not needs_yb and not needs_cond and not needs_agbg:
            continue
        if not comp.zpid:
            continue
        data = _api_get(PROPERTY_BY_ZPID_ENDPOINT, {"zpid": comp.zpid}, api_key)
        if not data:
            continue
        changed = False
        if needs_yb:
            yb = int(data.get("yearBuilt") or 0)
            gs = int(data.get("garageSpaces") or 0)
            if yb:
                comp.year_built = yb; changed = True
            if gs:
                comp.garage_spaces = gs; changed = True
        if needs_agbg:
            reso = data.get("resoFacts") or {}
            auto_ag = _safe_int(reso.get("aboveGradeFinishedArea"))
            auto_bg = _safe_int(reso.get("belowGradeFinishedArea"))
            if auto_ag and auto_bg:
                comp.ag_sqft = auto_ag
                comp.bg_finished_sqft = auto_bg
                changed = True
        if needs_cond:
            desc = data.get("description") or ""
            comp.description = desc
            auto = detect_condition_from_description(desc)
            if auto:
                comp.condition = auto
                comp.condition_source = "auto"
                changed = True
        if changed:
            enriched += 1
        time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))

    if enriched:
        logger.info("Enriched %d/%d comps (yearBuilt/garage/condition)", enriched, len(comps))
    return enriched


def apply_condition_overrides(comps: list[CompProperty], overrides: dict) -> int:
    """Apply zpid/address overrides to comps. Overrides beat auto-detect.

    Each override may carry condition, ag_sqft, and/or bg_sqft — applied
    independently. Returns count of comps touched.
    """
    if not overrides:
        return 0
    applied = 0
    for c in comps:
        key_zpid = f"zpid:{c.zpid}" if c.zpid else None
        key_addr = f"addr:{c.address.lower().strip()}" if c.address else None
        entry = None
        if key_zpid and key_zpid in overrides:
            entry = overrides[key_zpid]
        elif key_addr and key_addr in overrides:
            entry = overrides[key_addr]
        if not entry:
            continue
        touched = False
        if entry.get("condition"):
            c.condition = entry["condition"]
            c.condition_source = "override"
            touched = True
        if entry.get("ag_sqft"):
            c.ag_sqft = entry["ag_sqft"]
            touched = True
        if entry.get("bg_sqft"):
            c.bg_finished_sqft = entry["bg_sqft"]
            touched = True
        if touched:
            applied += 1
    if applied:
        logger.info("Applied %d override entry(s) from file", applied)
    return applied


def fetch_comparable_sales(subject: SubjectProperty, radius_miles: float = DEFAULT_RADIUS_MILES,
                           months_back: int = DEFAULT_MONTHS_BACK,
                           api_key: str = "",
                           condition_overrides: dict | None = None) -> list[CompProperty]:
    """Fetch comparable sold properties near the subject property.

    Two-source strategy:
      1. /similar-sale-homes — Zillow's own similarity engine. Best quality when it
         works, but 404s for properties with sparse Zillow records.
      2. /search?home_status=RECENTLY_SOLD — geographic ZIP-scoped fallback.
         Always available, paginated by date.
    Falls through 1 → 2 automatically. If the result is still thin, expands the
    radius/lookback and reruns. Condition overrides are applied to returned comps
    so later enrichment won't overwrite user-labeled values.
    """
    api_key = api_key or config.OPENWEBNINJA_API_KEY
    if not api_key:
        return []

    cutoff_date = (datetime.now() - timedelta(days=months_back * 30)).strftime("%Y-%m-%d")

    # Source 1: Zillow's similar-sale-homes
    full_address = f"{subject.address} {subject.city} {subject.state} {subject.zip_code}"
    data = _api_get(COMPS_ENDPOINT, {"address": full_address}, api_key)
    time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))

    items = data if isinstance(data, list) else (data.get("comps") or data.get("results") or []) if data else []
    comps = [c for c in (_item_to_comp(it, subject, radius_miles, cutoff_date) for it in items) if c]

    if comps:
        logger.info("Similar-sale-homes returned %d comps within %.1f mi (last %d months)",
                    len(comps), radius_miles, months_back)
    else:
        logger.info("Similar-sale-homes empty — falling back to ZIP sold search")
        comps = _fetch_sold_via_search(subject, radius_miles, cutoff_date, api_key)

    # Real expansion: if still thin, retry with widened bounds
    if len(comps) < MIN_COMPS and (radius_miles < MAX_RADIUS_MILES or months_back < MAX_MONTHS_BACK):
        new_radius = min(radius_miles * 1.5, MAX_RADIUS_MILES)
        new_months = min(months_back + 3, MAX_MONTHS_BACK)
        if new_radius > radius_miles or new_months > months_back:
            logger.info("Only %d comps — expanding to %.1f mi / %d months and refetching",
                        len(comps), new_radius, new_months)
            expanded_cutoff = (datetime.now() - timedelta(days=new_months * 30)).strftime("%Y-%m-%d")
            comps = _fetch_sold_via_search(subject, new_radius, expanded_cutoff, api_key)

    # Stamp in any user-labeled conditions before enrichment runs auto-detect
    if condition_overrides:
        apply_condition_overrides(comps, condition_overrides)

    return comps


# ── Similarity scoring ────────────────────────────────────────────────

def _score_similarity(subject: SubjectProperty, comp: CompProperty) -> float:
    """Score how similar a comp is to the subject (0.0 to 1.0, higher = more similar)."""
    score = 1.0
    penalties = []

    # Square footage (most important)
    if subject.sqft and comp.sqft:
        sqft_diff_pct = abs(subject.sqft - comp.sqft) / subject.sqft
        if sqft_diff_pct > 0.30:
            score -= 0.30
            penalties.append(f"sqft {sqft_diff_pct:.0%} diff")
        elif sqft_diff_pct > 0.20:
            score -= 0.15
        elif sqft_diff_pct > 0.10:
            score -= 0.05

    # Bedrooms
    if subject.bedrooms and comp.bedrooms:
        bed_diff = abs(subject.bedrooms - comp.bedrooms)
        if bed_diff > 2:
            score -= 0.25
        elif bed_diff > 1:
            score -= 0.10
        elif bed_diff == 1:
            score -= 0.03

    # Bathrooms
    if subject.bathrooms and comp.bathrooms:
        bath_diff = abs(subject.bathrooms - comp.bathrooms)
        if bath_diff > 2:
            score -= 0.20
        elif bath_diff > 1:
            score -= 0.08

    # Year built
    if subject.year_built and comp.year_built:
        age_diff = abs(subject.year_built - comp.year_built)
        if age_diff > 20:
            score -= 0.20
        elif age_diff > 10:
            score -= 0.08
        elif age_diff > 5:
            score -= 0.03

    # Distance (closer = better)
    if comp.distance_miles > 0.75:
        score -= 0.15
    elif comp.distance_miles > 0.5:
        score -= 0.08
    elif comp.distance_miles > 0.25:
        score -= 0.03

    # Property type mismatch
    if subject.property_type and comp.property_type:
        if subject.property_type.upper() != comp.property_type.upper():
            score -= 0.20

    # Recency bonus (sold more recently = better)
    if comp.sold_date:
        try:
            sold_dt = datetime.strptime(comp.sold_date[:10], "%Y-%m-%d")
            days_ago = (datetime.now() - sold_dt).days
            if days_ago < 30:
                score += 0.05
            elif days_ago < 90:
                score += 0.02
            elif days_ago > 180:
                score -= 0.05
        except ValueError:
            pass

    return max(0.0, min(1.0, score))


# ── Adjustment engine ─────────────────────────────────────────────────

def _derive_reno_premium(comps: list[CompProperty], subject_effective_sqft: float,
                         profile: dict) -> tuple[float, str, str]:
    """Derive the full-reno-premium from the actual comp pool when possible.

    Splits comps into as-is (condition_index ≤ 0.1) and full (≥ 0.85) buckets,
    requires ≥2 in each, and computes:
        observed = (median_full_ppsf - median_asis_ppsf) × subject_effective_sqft
    using effective sqft (AG + BG × bg_pct_of_ag). Falls back to the profile
    default if sample size is insufficient or the observed value is implausible
    (< 0.3× or > 3× the fallback). Returns (value, source, detail_msg).
    """
    fallback = profile["full_reno_premium"]
    if not comps or subject_effective_sqft <= 0:
        return fallback, "fallback", "no comps / no subject sqft"

    bg_pct = profile["bg_pct_of_ag"]

    def eff_sqft(c: CompProperty) -> float:
        return (c.ag_sqft or c.sqft) + c.bg_finished_sqft * bg_pct

    def ppsf(c: CompProperty) -> float:
        s = eff_sqft(c)
        return c.sold_price / s if s else 0.0

    asis = [c for c in comps if c.condition and normalize_condition(c.condition) <= 0.1]
    full = [c for c in comps if c.condition and normalize_condition(c.condition) >= 0.85]

    if len(asis) < 2 or len(full) < 2:
        return fallback, "fallback", f"need ≥2 in each bucket (got {len(asis)} as-is, {len(full)} full)"

    asis_ppsfs = sorted(ppsf(c) for c in asis if ppsf(c))
    full_ppsfs = sorted(ppsf(c) for c in full if ppsf(c))
    if not asis_ppsfs or not full_ppsfs:
        return fallback, "fallback", "empty PPSFs after filtering"

    median_asis = asis_ppsfs[len(asis_ppsfs) // 2]
    median_full = full_ppsfs[len(full_ppsfs) // 2]
    observed = (median_full - median_asis) * subject_effective_sqft

    if observed < fallback * 0.3 or observed > fallback * 3.0:
        return (fallback, "fallback",
                f"observed ${observed:,.0f} outside sanity range [0.3×, 3.0×] of fallback ${fallback:,.0f}")

    detail = (f"median full PPSF ${median_full:.2f} vs as-is ${median_asis:.2f} "
              f"× {subject_effective_sqft:.0f} eff sqft (n_full={len(full)}, n_asis={len(asis)})")
    return observed, "derived", detail


def _calculate_adjustments(subject: SubjectProperty, comp: CompProperty,
                           profile: dict | None = None) -> dict:
    """Calculate dollar adjustments from comp to subject property.

    Uses the regional adjustment profile (sqft/bd/ba/yr/lot/garage/market rates
    plus full_reno_premium). Condition adjustment converts comps to subject's
    target_condition equivalent via the CONDITION_INDEX scale.
    """
    p = profile or get_adj_profile(subject.state)
    adjustments = {}

    # Square footage: split into above-grade (full rate) and below-grade finished
    # (at bg_pct_of_ag of full rate). ag_sqft falls back to sqft when not split.
    subj_ag = subject.ag_sqft or subject.sqft
    comp_ag = comp.ag_sqft or comp.sqft
    if subj_ag and comp_ag:
        ag_diff = subj_ag - comp_ag
        if ag_diff != 0:
            adjustments["ag_sqft"] = round(ag_diff * p["sqft"])

    # BG finished adjustment — runs even if both sides are 0, since a difference
    # between subject having a basement and comp not having one is material.
    bg_diff = subject.bg_finished_sqft - comp.bg_finished_sqft
    if bg_diff != 0:
        adjustments["bg_sqft"] = round(bg_diff * p["sqft"] * p["bg_pct_of_ag"])

    # Bedroom adjustment
    if subject.bedrooms and comp.bedrooms:
        bed_diff = subject.bedrooms - comp.bedrooms
        if bed_diff != 0:
            adjustments["bedrooms"] = round(bed_diff * p["bedroom"])

    # Bathroom adjustment
    if subject.bathrooms and comp.bathrooms:
        bath_diff = subject.bathrooms - comp.bathrooms
        if bath_diff != 0:
            adjustments["bathrooms"] = round(bath_diff * p["bathroom"])

    # Age / year built adjustment
    if subject.year_built and comp.year_built:
        year_diff = subject.year_built - comp.year_built
        if year_diff != 0:
            adjustments["year_built"] = round(year_diff * p["year_built"])

    # Lot size adjustment (capped)
    if subject.lot_sqft and comp.lot_sqft:
        lot_diff = subject.lot_sqft - comp.lot_sqft
        if lot_diff != 0:
            adj = lot_diff * p["lot_sqft"]
            adj = max(-p["lot_max"], min(p["lot_max"], adj))
            adjustments["lot_size"] = round(adj)

    # Garage adjustment
    garage_diff = subject.garage_spaces - comp.garage_spaces
    if garage_diff != 0:
        adjustments["garage"] = round(garage_diff * p["garage"])

    # Condition adjustment — converts comp to subject's target-condition equivalent.
    # Positive = comp condition is below target, add premium to reach target.
    target_idx = normalize_condition(subject.target_condition or "full")
    comp_idx = normalize_condition(comp.condition) if comp.condition else 0.5
    cond_delta = target_idx - comp_idx
    if cond_delta != 0:
        adjustments["condition"] = round(cond_delta * p["full_reno_premium"])

    # Market conditions (time) adjustment
    if comp.sold_date:
        try:
            sold_dt = datetime.strptime(comp.sold_date[:10], "%Y-%m-%d")
            months_ago = (datetime.now() - sold_dt).days / 30.0
            if months_ago > 1:
                adj = comp.sold_price * p["market_pct_mo"] * months_ago
                adjustments["market_conditions"] = round(adj)
        except ValueError:
            pass

    return adjustments


def _apply_adjustments(comp: CompProperty, adjustments: dict) -> float:
    """Apply adjustments to comp's sold price and return adjusted price."""
    total_adj = sum(adjustments.values())
    return comp.sold_price + total_adj


# ── Two-Bucket classification ────────────────────────────────────────

def _classify_bucket(comp: CompProperty) -> str:
    """Classify comp into Bucket A or Bucket B.

    Bucket A: Non-disclosure baseline comps — properties with limited price
              transparency (typical in TN as a non-disclosure state).
              Uses Zillow/MLS-reported data as proxy.
    Bucket B: Disclosure/verified comps — properties with confirmed sale
              prices from MLS (listed and sold through agent).

    In practice for TN (non-disclosure state), all comps come through
    Zillow/MLS so we classify based on data completeness:
    - Bucket B: Has complete MLS data (sold through agent, DOM tracked)
    - Bucket A: Limited data (off-market sale, FSBO, etc.)
    """
    if comp.days_on_market > 0:
        return "B"  # Was listed on MLS — has disclosure data
    return "A"  # No DOM data — likely off-market/non-disclosure sale


# ── ARV calculation ───────────────────────────────────────────────────

def calculate_arv(subject: SubjectProperty, comps: list[CompProperty]) -> ARVResult:
    """Calculate After Repair Value using Two-Bucket methodology.

    1. Score and rank comps by similarity
    2. Classify into Bucket A (non-disclosure) and Bucket B (disclosure)
    3. Apply property-specific adjustments
    4. Calculate weighted ARV with confidence bands
    """
    if not comps:
        return ARVResult(confidence="none", confidence_reason="No comparable sales found")

    # Score and sort by similarity
    for comp in comps:
        comp.similarity_score = _score_similarity(subject, comp)

    comps.sort(key=lambda c: c.similarity_score, reverse=True)

    # Take top comps
    selected = comps[:MAX_COMPS]

    # Enrich missing yearBuilt/garageSpaces + auto-detect condition from descriptions
    enrich_comps_with_details(selected)

    # Shallow-copy profile so we can override full_reno_premium per-run without
    # mutating the module-level dict
    profile = dict(get_adj_profile(subject.state))

    # Self-calibrating renovation premium: derive from actual as-is vs full
    # comp bucket spread when we have enough samples. Falls back to profile
    # default otherwise. Uses full pool (not just selected) for larger N.
    subj_eff_sqft = (subject.ag_sqft or subject.sqft) + subject.bg_finished_sqft * profile["bg_pct_of_ag"]
    premium, premium_source, premium_detail = _derive_reno_premium(comps, subj_eff_sqft, profile)
    profile["full_reno_premium"] = premium
    logger.info("Reno premium (%s): $%s — %s", premium_source, f"{premium:,.0f}", premium_detail)

    # Classify buckets and calculate adjustments
    for comp in selected:
        comp.bucket = _classify_bucket(comp)
        comp.adjustments = _calculate_adjustments(subject, comp, profile)
        comp.adjusted_price = _apply_adjustments(comp, comp.adjustments)

    bucket_a = [c for c in selected if c.bucket == "A"]
    bucket_b = [c for c in selected if c.bucket == "B"]

    # Calculate PPSF from adjusted prices
    ppsf_values = []
    for comp in selected:
        if comp.sqft:
            ppsf_values.append(comp.adjusted_price / comp.sqft)

    # Weighted ARV: Bucket B gets 70% weight, Bucket A gets 30%
    # (disclosure data is more reliable)
    adj_prices = [c.adjusted_price for c in selected if c.adjusted_price > 0]

    if not adj_prices:
        return ARVResult(confidence="none", confidence_reason="No valid adjusted prices")

    if bucket_b:
        bucket_b_avg = sum(c.adjusted_price for c in bucket_b) / len(bucket_b)
    else:
        bucket_b_avg = 0

    if bucket_a:
        bucket_a_avg = sum(c.adjusted_price for c in bucket_a) / len(bucket_a)
    else:
        bucket_a_avg = 0

    # Weighted average
    if bucket_b_avg and bucket_a_avg:
        arv_mid = bucket_b_avg * 0.70 + bucket_a_avg * 0.30
    elif bucket_b_avg:
        arv_mid = bucket_b_avg
    else:
        arv_mid = bucket_a_avg

    # Market sentiment layer: apply phase-based multiplier to arv_mid AFTER the
    # per-comp time-appreciation adjustments already baked in. This captures
    # current buyer behavior (DOM-based) not captured by historical appreciation.
    arv_mid_pre_sentiment = arv_mid
    market_phase, market_dom_avg, sentiment_mult = classify_market_phase(selected)
    if sentiment_mult != 0:
        arv_mid = arv_mid * (1 + sentiment_mult)
        logger.info("Market sentiment: %s (avg DOM %d days) → %+.0f%% applied to ARV mid",
                    market_phase, market_dom_avg, sentiment_mult * 100)

    # Confidence bands: tighter for disclosure states (hard sold data),
    # wider for non-disclosure states (sold prices are estimated).
    spread = max(adj_prices) - min(adj_prices)
    spread_pct = (spread / arv_mid * 100) if arv_mid else 0
    state_type = profile.get("state_type", "non-disclosure")

    # (low_mult, high_mult) by spread tier, per state type
    if state_type == "disclosure":
        if spread_pct < 10:
            low_mult, high_mult, confidence = 0.97, 1.03, "high"
            confidence_reason = f"Tight comp spread ({spread_pct:.0f}%), {len(selected)} comps [disclosure]"
        elif spread_pct < 20:
            low_mult, high_mult, confidence = 0.93, 1.05, "medium"
            confidence_reason = f"Moderate comp spread ({spread_pct:.0f}%), {len(selected)} comps [disclosure]"
        else:
            low_mult, high_mult, confidence = 0.88, 1.08, "low"
            confidence_reason = f"Wide comp spread ({spread_pct:.0f}%) [disclosure] — verify with local knowledge"
    else:  # non-disclosure — sold prices estimated, wider bands
        if spread_pct < 10:
            low_mult, high_mult, confidence = 0.95, 1.05, "high"
            confidence_reason = f"Tight comp spread ({spread_pct:.0f}%), {len(selected)} comps [non-disclosure]"
        elif spread_pct < 20:
            low_mult, high_mult, confidence = 0.90, 1.08, "medium"
            confidence_reason = f"Moderate comp spread ({spread_pct:.0f}%), {len(selected)} comps [non-disclosure]"
        else:
            low_mult, high_mult, confidence = 0.85, 1.10, "low"
            confidence_reason = f"Wide comp spread ({spread_pct:.0f}%) [non-disclosure] — verify with local knowledge"

    arv_low = arv_mid * low_mult
    arv_high = arv_mid * high_mult

    # Fewer comps = lower confidence
    if len(selected) < MIN_COMPS:
        confidence = "low"
        confidence_reason = f"Only {len(selected)} comps found (minimum {MIN_COMPS} recommended)"

    avg_adj = sum(abs(sum(c.adjustments.values())) for c in selected) / len(selected) if selected else 0

    return ARVResult(
        arv_low=round(arv_low),
        arv_mid=round(arv_mid),
        arv_high=round(arv_high),
        arv_mid_pre_sentiment=round(arv_mid_pre_sentiment),
        confidence=confidence,
        confidence_reason=confidence_reason,
        ppsf_avg=round(sum(ppsf_values) / len(ppsf_values), 2) if ppsf_values else 0,
        ppsf_range=(round(min(ppsf_values), 2), round(max(ppsf_values), 2)) if ppsf_values else (0, 0),
        comp_count=len(selected),
        bucket_a_count=len(bucket_a),
        bucket_b_count=len(bucket_b),
        avg_adjustment=round(avg_adj),
        spread_pct=round(spread_pct, 1),
        market_phase=market_phase,
        market_dom_avg=market_dom_avg,
        sentiment_adj_pct=sentiment_mult,
        reno_premium_used=round(premium),
        reno_premium_source=premium_source,
    )


# ── Excel report generation ──────────────────────────────────────────

# Styles
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_MONEY_FMT = '#,##0'
_PCT_FMT = '0.0%'
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_VALUE_FONT = Font(name="Calibri", bold=True, size=13, color="222222")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT = Font(name="Calibri", bold=True, color="006100")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_column_widths(ws, min_width: int = 12, max_width: int = 35) -> None:
    """Auto-size columns based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _fmt_money(value: float) -> str:
    """Format number as currency string."""
    if not value:
        return "$0"
    return f"${value:,.0f}"


def generate_comp_report(subject: SubjectProperty, comps: list[CompProperty],
                         arv: ARVResult, output_path: str) -> str:
    """Generate a 7-tab Excel workbook comp report.

    Tabs:
    1. Executive Summary — subject property + ARV range
    2. Subject Property — full detail
    3. Comparable Sales — all comps with distance, date, similarity
    4. Adjustments Detail — per-comp adjustment breakdown
    5. Market Analysis — PPSF trends, DOM, market direction
    6. ARV Calculation — Two-Bucket weighted result with confidence bands
    7. Sources & Notes
    """
    wb = Workbook()

    # ── Tab 1: Executive Summary ──────────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"

    ws.cell(row=1, column=1, value="Comp Analysis Report").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"{subject.address}, {subject.city}, {subject.state} {subject.zip_code}").font = _SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = _LABEL_FONT

    row = 5
    summary_data = [
        ("ARV (Low)", _fmt_money(arv.arv_low)),
        ("ARV (Mid — Recommended)", _fmt_money(arv.arv_mid)),
        ("ARV (High)", _fmt_money(arv.arv_high)),
        ("", ""),
        ("Confidence Level", arv.confidence.upper()),
        ("Confidence Reason", arv.confidence_reason),
        ("", ""),
        ("Avg PPSF", f"${arv.ppsf_avg:,.2f}"),
        ("PPSF Range", f"${arv.ppsf_range[0]:,.2f} — ${arv.ppsf_range[1]:,.2f}"),
        ("Comps Analyzed", str(arv.comp_count)),
        ("Bucket A (Non-Disclosure)", str(arv.bucket_a_count)),
        ("Bucket B (Disclosure/MLS)", str(arv.bucket_b_count)),
        ("Avg Gross Adjustment", _fmt_money(arv.avg_adjustment)),
        ("Comp Spread", f"{arv.spread_pct:.1f}%"),
        ("", ""),
        ("Subject Zestimate", _fmt_money(subject.zestimate)),
        ("Subject Property Type", subject.property_type),
        ("Subject Sqft", f"{subject.sqft:,}" if subject.sqft else "N/A"),
        ("Subject Bed/Bath", f"{subject.bedrooms}bd / {subject.bathrooms}ba"),
        ("Subject Year Built", str(subject.year_built) if subject.year_built else "N/A"),
    ]
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = _VALUE_FONT
        if label == "Confidence Level":
            if arv.confidence == "high":
                cell.fill = _GREEN_FILL
                cell.font = _GREEN_FONT
            elif arv.confidence == "medium":
                cell.fill = _YELLOW_FILL
            else:
                cell.fill = _RED_FILL
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35

    # ── Tab 2: Subject Property ──────────────────────────────────────
    ws2 = wb.create_sheet("Subject Property")
    ws2.cell(row=1, column=1, value="Subject Property Details").font = _TITLE_FONT

    props = [
        ("Address", subject.address),
        ("City", subject.city),
        ("State", subject.state),
        ("ZIP", subject.zip_code),
        ("Latitude", str(subject.latitude) if subject.latitude else ""),
        ("Longitude", str(subject.longitude) if subject.longitude else ""),
        ("Property Type", subject.property_type),
        ("Total Finished Sqft", f"{subject.sqft:,}" if subject.sqft else ""),
        ("Above-Grade Sqft", f"{subject.ag_sqft:,}" if subject.ag_sqft else ""),
        ("Below-Grade Finished Sqft", f"{subject.bg_finished_sqft:,}" if subject.bg_finished_sqft else "0"),
        ("Bedrooms", str(subject.bedrooms)),
        ("Bathrooms", str(subject.bathrooms)),
        ("Year Built", str(subject.year_built) if subject.year_built else ""),
        ("Lot Size (sqft)", f"{subject.lot_sqft:,}" if subject.lot_sqft else ""),
        ("Garage Spaces", str(subject.garage_spaces)),
        ("Target Condition (after rehab)", subject.target_condition),
        ("Zestimate", _fmt_money(subject.zestimate)),
        ("MLS Status", subject.mls_status),
        ("Last Sold Date", subject.last_sold_date),
        ("Last Sold Price", _fmt_money(subject.last_sold_price)),
    ]
    for i, (label, value) in enumerate(props, 3):
        ws2.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws2.cell(row=i, column=2, value=value).font = _VALUE_FONT
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 30

    # ── Tab 3: Comparable Sales ──────────────────────────────────────
    ws3 = wb.create_sheet("Comparable Sales")
    ws3.cell(row=1, column=1, value="Comparable Sales").font = _TITLE_FONT

    comp_headers = ["#", "Address", "City", "ZIP", "Distance (mi)", "Sold Price",
                    "Sold Date", "Total Sqft", "AG Sqft", "BG Sqft",
                    "Bed", "Bath", "Year Built", "PPSF",
                    "Condition", "Cond Source", "Similarity", "Bucket", "Adjusted Price"]
    _write_header_row(ws3, 3, comp_headers)

    for i, comp in enumerate(comps[:MAX_COMPS], 1):
        row = i + 3
        values = [
            i, comp.address, comp.city, comp.zip_code,
            comp.distance_miles, comp.sold_price, comp.sold_date,
            comp.sqft, comp.ag_sqft, comp.bg_finished_sqft,
            comp.bedrooms, comp.bathrooms, comp.year_built,
            round(comp.ppsf, 2),
            comp.condition or "(unknown)",
            comp.condition_source or "—",
            f"{comp.similarity_score:.0%}",
            comp.bucket, comp.adjusted_price,
        ]
        for col, val in enumerate(values, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            if col in (6, 19):  # Money columns (Sold Price, Adjusted Price)
                cell.number_format = _MONEY_FMT
            cell.border = _THIN_BORDER

    _auto_column_widths(ws3)

    # ── Tab 4: Adjustments Detail ────────────────────────────────────
    ws4 = wb.create_sheet("Adjustments Detail")
    ws4.cell(row=1, column=1, value="Per-Comp Adjustment Breakdown").font = _TITLE_FONT

    adj_types = ["ag_sqft", "bg_sqft", "bedrooms", "bathrooms", "year_built", "lot_size", "garage", "condition", "market_conditions"]
    adj_headers = ["Comp #", "Address", "Sold Price"] + [a.replace("_", " ").title() for a in adj_types] + ["Total Adj", "Adjusted Price"]
    _write_header_row(ws4, 3, adj_headers)

    for i, comp in enumerate(comps[:MAX_COMPS], 1):
        row = i + 3
        values = [i, comp.address, comp.sold_price]
        total_adj = 0
        for adj_type in adj_types:
            adj_val = comp.adjustments.get(adj_type, 0)
            total_adj += adj_val
            values.append(adj_val)
        values.append(total_adj)
        values.append(comp.adjusted_price)
        for col, val in enumerate(values, 1):
            cell = ws4.cell(row=row, column=col, value=val)
            if col >= 3:
                cell.number_format = _MONEY_FMT
            cell.border = _THIN_BORDER

    _auto_column_widths(ws4)

    # ── Tab 5: Market Analysis ───────────────────────────────────────
    ws5 = wb.create_sheet("Market Analysis")
    ws5.cell(row=1, column=1, value="Market Analysis").font = _TITLE_FONT

    # PPSF analysis
    ws5.cell(row=3, column=1, value="Price Per Square Foot Analysis").font = _SUBTITLE_FONT
    ppsf_data = [
        ("Average PPSF", f"${arv.ppsf_avg:,.2f}"),
        ("PPSF Range", f"${arv.ppsf_range[0]:,.2f} — ${arv.ppsf_range[1]:,.2f}"),
        ("Subject Implied Value (Avg PPSF)", _fmt_money(arv.ppsf_avg * subject.sqft) if subject.sqft else "N/A"),
    ]
    for i, (label, value) in enumerate(ppsf_data, 4):
        ws5.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws5.cell(row=i, column=2, value=value).font = _VALUE_FONT

    # Days on market
    ws5.cell(row=8, column=1, value="Days on Market Analysis").font = _SUBTITLE_FONT
    dom_values = [c.days_on_market for c in comps[:MAX_COMPS] if c.days_on_market > 0]
    if dom_values:
        ws5.cell(row=9, column=1, value="Average DOM").font = _LABEL_FONT
        ws5.cell(row=9, column=2, value=f"{sum(dom_values) / len(dom_values):.0f} days").font = _VALUE_FONT
        ws5.cell(row=10, column=1, value="Median DOM").font = _LABEL_FONT
        sorted_dom = sorted(dom_values)
        median_dom = sorted_dom[len(sorted_dom) // 2]
        ws5.cell(row=10, column=2, value=f"{median_dom} days").font = _VALUE_FONT

    # Market direction
    profile = get_adj_profile(subject.state)
    ws5.cell(row=12, column=1, value="Market Direction").font = _SUBTITLE_FONT
    ws5.cell(row=13, column=1, value="Monthly Appreciation Rate").font = _LABEL_FONT
    ws5.cell(row=13, column=2, value=f"{profile['market_pct_mo'] * 100:.1f}%").font = _VALUE_FONT
    ws5.cell(row=14, column=1, value="Annualized Appreciation").font = _LABEL_FONT
    ws5.cell(row=14, column=2, value=f"{profile['market_pct_mo'] * 12 * 100:.1f}%").font = _VALUE_FONT

    ws5.column_dimensions["A"].width = 35
    ws5.column_dimensions["B"].width = 25

    # ── Tab 6: ARV Calculation ───────────────────────────────────────
    ws6 = wb.create_sheet("ARV Calculation")
    ws6.cell(row=1, column=1, value="Two-Bucket ARV Calculation").font = _TITLE_FONT

    ws6.cell(row=3, column=1, value="Methodology").font = _SUBTITLE_FONT
    ws6.cell(row=4, column=1, value="Bucket A (Non-Disclosure): 30% weight — Off-market/FSBO sales with limited price transparency").font = _LABEL_FONT
    ws6.cell(row=5, column=1, value="Bucket B (Disclosure/MLS): 70% weight — Agent-listed sales with confirmed pricing").font = _LABEL_FONT
    ws6.cell(row=6, column=1, value=profile["disclosure_note"]).font = _LABEL_FONT

    ws6.cell(row=8, column=1, value="Bucket A Comps").font = _SUBTITLE_FONT
    bucket_a = [c for c in comps[:MAX_COMPS] if c.bucket == "A"]
    bucket_b = [c for c in comps[:MAX_COMPS] if c.bucket == "B"]
    if bucket_a:
        avg_a = sum(c.adjusted_price for c in bucket_a) / len(bucket_a)
        ws6.cell(row=9, column=1, value=f"Count: {len(bucket_a)}  |  Avg Adjusted: {_fmt_money(avg_a)}").font = _VALUE_FONT
    else:
        ws6.cell(row=9, column=1, value="No Bucket A comps").font = _LABEL_FONT

    ws6.cell(row=11, column=1, value="Bucket B Comps").font = _SUBTITLE_FONT
    if bucket_b:
        avg_b = sum(c.adjusted_price for c in bucket_b) / len(bucket_b)
        ws6.cell(row=12, column=1, value=f"Count: {len(bucket_b)}  |  Avg Adjusted: {_fmt_money(avg_b)}").font = _VALUE_FONT
    else:
        ws6.cell(row=12, column=1, value="No Bucket B comps").font = _LABEL_FONT

    ws6.cell(row=14, column=1, value="Final ARV").font = _SUBTITLE_FONT
    arv_display = [
        ("ARV Low (Conservative)", _fmt_money(arv.arv_low)),
        ("ARV Mid (Recommended)", _fmt_money(arv.arv_mid)),
        ("ARV High (Optimistic)", _fmt_money(arv.arv_high)),
        ("Confidence", arv.confidence.upper()),
        ("Reason", arv.confidence_reason),
    ]
    for i, (label, value) in enumerate(arv_display, 15):
        ws6.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws6.cell(row=i, column=2, value=value).font = _VALUE_FONT

    ws6.column_dimensions["A"].width = 40
    ws6.column_dimensions["B"].width = 30

    # ── Tab 7: Sources & Notes ───────────────────────────────────────
    ws7 = wb.create_sheet("Sources & Notes")
    ws7.cell(row=1, column=1, value="Sources & Notes").font = _TITLE_FONT

    notes = [
        "Data Source: OpenWeb Ninja Real-Time Zillow Data API",
        "Comparable sales sourced from Zillow's similar-sale-homes + recently-sold search endpoints",
        "",
        f"Adjustment Methodology ({profile['region_label']} calibration):",
        f"  Above-Grade Sqft: ${profile['sqft']:,.0f} per sqft difference",
        f"  Below-Grade Finished: ${profile['sqft'] * profile['bg_pct_of_ag']:,.0f} per sqft "
        f"({profile['bg_pct_of_ag'] * 100:.0f}% of AG rate)",
        f"  Bedrooms: ${profile['bedroom']:,.0f} per bedroom difference",
        f"  Bathrooms: ${profile['bathroom']:,.0f} per bathroom difference",
        f"  Year Built: ${profile['year_built']:,.0f} per year difference",
        f"  Lot Size: ${profile['lot_sqft']:,.2f} per sqft (capped at ${profile['lot_max']:,.0f})",
        f"  Garage: ${profile['garage']:,.0f} per stall difference",
        f"  Market Conditions: {profile['market_pct_mo'] * 100:.1f}% per month appreciation",
        f"  Condition (full-reno premium): ${profile['full_reno_premium']:,.0f} (as-is → full)",
        f"    Target subject condition: {subject.target_condition}",
        "",
        "Two-Bucket Weighting:",
        "  Bucket A (Non-Disclosure): 30% weight",
        "  Bucket B (Disclosure/MLS): 70% weight",
        "",
        "Confidence Bands:",
        "  High (<10% spread): ±5% of mid ARV",
        "  Medium (10-20% spread): -10%/+8% of mid ARV",
        "  Low (>20% spread): -15%/+10% of mid ARV",
        "",
        "Conservative bias: low-end ARV is intentionally wider.",
        "A high ARV that doesn't hold up kills your deal.",
        "A conservative ARV that comes in low leaves room for upside.",
        "",
        f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Region: {profile['region_label']}",
    ]
    for i, note in enumerate(notes, 3):
        ws7.cell(row=i, column=1, value=note).font = _LABEL_FONT

    ws7.column_dimensions["A"].width = 70

    # Save
    wb.save(output_path)
    logger.info("Comp report saved to %s", output_path)
    return output_path


# ── PDF report generation ────────────────────────────────────────────

def generate_comp_pdf(subject: SubjectProperty, comps: list[CompProperty],
                      arv: ARVResult, output_path: str) -> str:
    """Generate a concise single-file PDF summary of the comp analysis.

    Produces a shareable deliverable with the headline ARV, subject specs,
    comp table, ARV math breakdown, and methodology notes. Complements the
    multi-tab Excel workbook for clients/partners who want one clean doc.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak, HRFlowable,
    )

    BRAND = colors.HexColor("#2F5496")
    LIGHT = colors.HexColor("#F0F8FF")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=20,
                                  textColor=BRAND, alignment=1, spaceAfter=6)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=11,
                                alignment=1, spaceAfter=12)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], fontSize=13,
                                    textColor=BRAND, spaceAfter=6, spaceBefore=10)
    label_style = ParagraphStyle("Label", parent=styles["Normal"], fontSize=10,
                                  leading=14)
    note_style = ParagraphStyle("Note", parent=styles["Normal"], fontSize=9,
                                 textColor=colors.grey, leading=11)
    arv_box_style = ParagraphStyle("ARVBox", parent=styles["Normal"], fontSize=24,
                                    textColor=BRAND, alignment=1, leading=30)

    doc = SimpleDocTemplate(output_path, pagesize=letter,
                             leftMargin=0.6*inch, rightMargin=0.6*inch,
                             topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    profile = get_adj_profile(subject.state)

    # ── Title + address + ARV headline ──
    story.append(Paragraph("PROPERTY VALUATION REPORT", title_style))
    story.append(Paragraph(
        f"{subject.address}, {subject.city}, {subject.state} {subject.zip_code}",
        sub_style,
    ))

    arv_table = Table(
        [[Paragraph("<b>ESTIMATED AFTER-REPAIR VALUE</b>", label_style)],
         [Paragraph(f"<b>{_fmt_money(arv.arv_mid)}</b>", arv_box_style)],
         [Paragraph(
             f"Range: {_fmt_money(arv.arv_low)} — {_fmt_money(arv.arv_high)}  "
             f"|  Confidence: <b>{arv.confidence.upper()}</b>",
             ParagraphStyle("ARVSub", parent=styles["Normal"], fontSize=10, alignment=1, textColor=colors.grey),
         )]],
        colWidths=[6.3*inch],
    )
    arv_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("BOX", (0, 0), (-1, -1), 1.5, BRAND),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(arv_table)
    story.append(Spacer(1, 14))

    # ── Subject Property ──
    story.append(Paragraph("SUBJECT PROPERTY", section_style))
    story.append(HRFlowable(width="100%", thickness=0.5, color=BRAND))
    subject_rows = [
        ["Address", subject.address],
        ["City / State / ZIP", f"{subject.city}, {subject.state} {subject.zip_code}"],
        ["Property Type", subject.property_type or "—"],
        ["Total Finished Sqft", f"{subject.sqft:,}" if subject.sqft else "—"],
        ["Above-Grade Sqft", f"{subject.ag_sqft:,}" if subject.ag_sqft else "—"],
        ["Below-Grade Finished", f"{subject.bg_finished_sqft:,}" if subject.bg_finished_sqft else "0"],
        ["Beds / Baths", f"{subject.bedrooms} / {subject.bathrooms}"],
        ["Year Built", str(subject.year_built) if subject.year_built else "—"],
        ["Lot Size", f"{subject.lot_sqft:,} sqft" if subject.lot_sqft else "—"],
        ["Garage", str(subject.garage_spaces)],
        ["Target Condition", subject.target_condition],
        ["Zestimate", _fmt_money(subject.zestimate)],
    ]
    subject_table = Table(subject_rows, colWidths=[2.2*inch, 4.1*inch])
    subject_table.setStyle(TableStyle([
        ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 10),
        ("FONT", (1, 0), (1, -1), "Helvetica", 10),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(subject_table)
    story.append(Spacer(1, 12))

    # ── Comparable Sales Table ──
    story.append(Paragraph("COMPARABLE SALES", section_style))
    story.append(HRFlowable(width="100%", thickness=0.5, color=BRAND))
    comp_header = ["#", "Address", "AG", "BG", "Bd/Ba", "Yr", "Sold", "Cond", "Adj Price"]
    comp_rows = [comp_header]
    for i, c in enumerate(comps[:MAX_COMPS], 1):
        comp_rows.append([
            str(i),
            c.address[:24],
            f"{c.ag_sqft:,}",
            f"{c.bg_finished_sqft:,}" if c.bg_finished_sqft else "0",
            f"{c.bedrooms}/{c.bathrooms}",
            str(c.year_built) if c.year_built else "—",
            _fmt_money(c.sold_price),
            (c.condition or "—")[:8],
            _fmt_money(c.adjusted_price),
        ])
    comp_table = Table(comp_rows, colWidths=[0.3*inch, 1.6*inch, 0.55*inch, 0.5*inch,
                                               0.6*inch, 0.45*inch, 0.9*inch, 0.65*inch, 0.95*inch])
    comp_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(comp_table)
    story.append(Spacer(1, 12))

    # ── ARV Calculation Breakdown ──
    story.append(Paragraph("ARV CALCULATION", section_style))
    story.append(HRFlowable(width="100%", thickness=0.5, color=BRAND))
    sentiment_line = f"{arv.sentiment_adj_pct:+.0%}" if arv.sentiment_adj_pct else "0%"
    reno_tag = f"(derived)" if arv.reno_premium_source == "derived" else "(profile default)"
    arv_rows = [
        ["Comp count selected", str(arv.comp_count)],
        ["Bucket A (off-market) / B (MLS)", f"{arv.bucket_a_count} / {arv.bucket_b_count}"],
        ["Bucket-weighted avg (pre-sentiment)", _fmt_money(arv.arv_mid_pre_sentiment)],
        ["Market phase", f"{arv.market_phase.replace('_', ' ').title()} (avg DOM {arv.market_dom_avg} days)"],
        ["Sentiment adjustment", sentiment_line],
        ["ARV mid (post-sentiment)", _fmt_money(arv.arv_mid)],
        ["ARV low (-{:.0%})".format(1 - arv.arv_low/arv.arv_mid if arv.arv_mid else 0),
         _fmt_money(arv.arv_low)],
        ["ARV high (+{:.0%})".format(arv.arv_high/arv.arv_mid - 1 if arv.arv_mid else 0),
         _fmt_money(arv.arv_high)],
        ["Comp spread", f"{arv.spread_pct:.1f}%"],
        ["Confidence", f"{arv.confidence.upper()} — {arv.confidence_reason}"],
        ["Reno premium used", f"{_fmt_money(arv.reno_premium_used)} {reno_tag}"],
    ]
    arv_table = Table(arv_rows, colWidths=[2.7*inch, 3.6*inch])
    arv_table.setStyle(TableStyle([
        ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 10),
        ("FONT", (1, 0), (1, -1), "Helvetica", 10),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("BACKGROUND", (0, 5), (-1, 5), LIGHT),  # Highlight final ARV row
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(arv_table)
    story.append(Spacer(1, 12))

    # ── Methodology Notes ──
    story.append(Paragraph("METHODOLOGY", section_style))
    story.append(HRFlowable(width="100%", thickness=0.5, color=BRAND))
    notes = [
        f"Calibration: {profile['region_label']} ({profile.get('state_type', 'disclosure')} state)",
        f"Adjustments: AG sqft ${profile['sqft']:,.0f}/sf | BG finished "
        f"${profile['sqft'] * profile['bg_pct_of_ag']:,.0f}/sf ({profile['bg_pct_of_ag']*100:.0f}% of AG) | "
        f"Bedroom ${profile['bedroom']:,.0f} | Bath ${profile['bathroom']:,.0f} | "
        f"Garage ${profile['garage']:,.0f} | Year-built ${profile['year_built']:,.0f}/yr | "
        f"Lot ${profile['lot_sqft']:,.2f}/sf (cap ${profile['lot_max']:,.0f})",
        f"Full-reno premium: {_fmt_money(arv.reno_premium_used)} (as-is → full), scaled linearly across condition tiers",
        f"Market appreciation (time): {profile['market_pct_mo']*100:.1f}% per month applied to each comp's sold date",
        f"Market sentiment (phase): {arv.sentiment_adj_pct:+.0%} based on avg sold-comp DOM of {arv.market_dom_avg} days",
        "Two-Bucket Weighting: Bucket B (MLS-disclosed) 70% / Bucket A (off-market) 30%",
        "Confidence bands scale with comp spread and state type (disclosure vs non-disclosure)",
        f"Data source: OpenWeb Ninja Real-Time Zillow Data API",
    ]
    for note in notes:
        story.append(Paragraph(f"• {note}", note_style))

    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"<i>Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Not a formal appraisal — use as an investment-decision benchmark</i>",
        note_style,
    ))

    doc.build(story)
    logger.info("Comp PDF saved to %s", output_path)
    return output_path


# ── Main entry point ──────────────────────────────────────────────────

DEFAULT_CONDITION_FILE = "data/condition_overrides.csv"


def run_comp_analysis(address: str, city: str = "", state: str = "TN",
                      zip_code: str = "", radius: float = DEFAULT_RADIUS_MILES,
                      months: int = DEFAULT_MONTHS_BACK,
                      output_path: str = "",
                      subject_overrides: dict | None = None,
                      target_condition: str = "full",
                      condition_file: str = DEFAULT_CONDITION_FILE) -> dict:
    """Run a full comp analysis for a property and generate the report.

    subject_overrides: dict of Zillow corrections (beds/baths/sqft/year_built/
        lot_sqft/garage/property_type). Any key provided wins over Zillow data.
    target_condition: condition label the subject will be valued at after rehab
        ("full" by default). Comps are adjusted up/down to match.
    condition_file: CSV of zpid/address → condition labels. Overrides win over
        auto-detection from Zillow's description field.

    Returns a dict with ARV results and the output file path.
    """
    logger.info("Starting comp analysis for: %s %s %s %s", address, city, state, zip_code)

    # Step 1: Fetch subject property details (with overrides)
    subject = fetch_subject_property(address, city, state, zip_code,
                                     overrides=subject_overrides,
                                     target_condition=target_condition)
    if not subject:
        logger.error("Could not fetch subject property data")
        return {"error": "Could not fetch subject property data"}

    profile = get_adj_profile(subject.state)
    logger.info("Subject: %s — %s sqft, %dbd/%sba, built %s, Zestimate %s | calibration: %s | target: %s",
                subject.address, f"{subject.sqft:,}" if subject.sqft else "?",
                subject.bedrooms, subject.bathrooms,
                subject.year_built or "?", _fmt_money(subject.zestimate),
                profile["region_label"], subject.target_condition)

    # Step 2: Load condition overrides and fetch comparable sales
    overrides = load_condition_overrides(condition_file)
    if overrides:
        logger.info("Loaded %d condition override entries from %s", len(overrides), condition_file)
    comps = fetch_comparable_sales(subject, radius, months, condition_overrides=overrides)
    if not comps:
        logger.warning("No comparable sales found — try expanding radius or time window")
        return {"error": "No comparable sales found", "subject": subject}

    # Step 3: Calculate ARV
    arv = calculate_arv(subject, comps)
    logger.info("ARV: %s (low) / %s (mid) / %s (high) — %s confidence",
                _fmt_money(arv.arv_low), _fmt_money(arv.arv_mid),
                _fmt_money(arv.arv_high), arv.confidence)

    # Step 4: Generate reports (Excel detail + PDF summary)
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_address = "".join(c if c.isalnum() or c in " -" else "_" for c in address)[:40]
        output_path = str(config.OUTPUT_DIR / f"comp_report_{safe_address}_{timestamp}.xlsx")

    report_path = generate_comp_report(subject, comps, arv, output_path)
    pdf_path = output_path.rsplit(".", 1)[0] + ".pdf"
    try:
        generate_comp_pdf(subject, comps, arv, pdf_path)
    except Exception as e:
        logger.warning("PDF generation failed: %s (Excel report still produced)", e)
        pdf_path = ""

    return {
        "subject": subject,
        "comps": comps[:MAX_COMPS],
        "arv": arv,
        "report_path": report_path,
        "pdf_path": pdf_path,
    }

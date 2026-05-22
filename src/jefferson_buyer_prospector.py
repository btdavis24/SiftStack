"""Jefferson County KY cash buyer prospector.

Pulls Investor Transactions from DataSift's Sold Properties surface for
Jefferson County across a configurable month range, aggregates by buyer
name, and ranks. Writes a buyer-prospecting Excel workbook plus a
DataSift-ready CSV.

This is the Phase 1 build — DataSift only. The optional Phase 1B (cross-
reference against Jefferson County deed records via JCD) is intentionally
not wired in yet so we can validate the DataSift-side ranking first.

Usage:
  python src/main.py buyer-prospect-jefferson --start 2026-03 --end 2026-03
  python src/main.py buyer-prospect-jefferson --months-back 12
"""

from __future__ import annotations

import asyncio
import csv
import logging
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config
from datasift_sold_properties import (
    SoldProperty, scrape_sold_properties, export_sold_csv,
)
from jefferson_buyer_cross_ref import (
    CrossRefResult, DeedOnlyBuyer, cross_reference,
)
from jefferson_deeds_scraper import (
    DeedTransfer, export_deed_transfers_csv, scrape_jefferson_deed_transfers,
)

logger = logging.getLogger(__name__)


# ── Buyer ranking ─────────────────────────────────────────────────────


@dataclass
class BuyerRanking:
    """Aggregated stats for one buyer across the scraped window."""
    buyer_name: str = ""
    transaction_count: int = 0
    total_invested: int = 0
    avg_purchase_price: int = 0
    min_purchase_price: int = 0
    max_purchase_price: int = 0
    months_active: int = 0
    first_month: str = ""
    last_month: str = ""
    is_entity: bool = False
    properties: list[str] = field(default_factory=list)
    score: float = 0.0
    rank: int = 0
    # Builder/bulk-acquirer detection — see _classify_buyers().
    # category ∈ {"flipper", "builder_bulk", "individual"}.
    # We rank flippers + individuals together (the wholesale-target audience)
    # and surface builder_bulk separately in their own Excel tab so they
    # don't pollute the main scorecard.
    category: str = "flipper"
    bulk_signal: str = ""           # human-readable reason if builder_bulk
    # JCD deed cross-reference (Phase 1B). 0 = unverified (DataSift flagged
    # them but no matching deed found in the public record).
    deed_verified_count: int = 0
    deed_first_filing: str = ""     # YYYY-MM-DD
    deed_last_filing: str = ""      # YYYY-MM-DD


# Entity keywords — copied & extended from buyer_prospector.py so we mark
# LLCs / Trusts / Corps consistently with the rest of the platform.
_ENTITY_KEYWORDS = (
    "LLC", "L L C", "CORP", "INC", "TRUST", "LP", "LLP", "LTD",
    "PROPERTIES", "HOLDINGS", "INVESTMENTS", "CAPITAL", "REALTY",
    "REAL ESTATE", "VENTURES", "LAND TRUST", "PARTNERS", "GROUP",
    "ASSOCIATES", "ENTERPRISES", "DEVELOPMENT", "FUND",
)

# Buyers excluded from the wholesale-target list. State agencies and
# non-profits acquire properties for affordable housing / rehab grants —
# they're not someone we'd ever wholesale to. Substring match on the
# uppercased buyer name. Add new entries here as they surface; keep this
# list small and high-confidence.
_EXCLUDED_BUYER_PATTERNS = (
    "HABITAT FOR HUMANITY",
    "KENTUCKY HOUSING",
    "LANDBANK AUTHORITY",       # Louisville & Jefferson County Landbank Authority Inc
)


def _is_excluded(name: str) -> bool:
    upper = (name or "").upper()
    return any(pat in upper for pat in _EXCLUDED_BUYER_PATTERNS)


def _is_entity(name: str) -> bool:
    upper = (name or "").upper()
    return any(kw in upper for kw in _ENTITY_KEYWORDS)


def aggregate_buyers(rows: list[SoldProperty]) -> list[BuyerRanking]:
    """Roll up SoldProperty rows by buyer name into BuyerRanking rows."""
    by_buyer: dict[str, list[SoldProperty]] = defaultdict(list)
    for r in rows:
        name = (r.buyer_name or "").strip().upper()
        if not name or len(name) < 3:
            continue
        by_buyer[name].append(r)

    rankings: list[BuyerRanking] = []
    for name, recs in by_buyer.items():
        prices = [r.sale_amount for r in recs if r.sale_amount > 0]
        months = sorted({r.sale_month for r in recs if r.sale_month})
        rankings.append(BuyerRanking(
            buyer_name=name,
            transaction_count=len(recs),
            total_invested=sum(prices),
            avg_purchase_price=round(sum(prices) / len(prices)) if prices else 0,
            min_purchase_price=min(prices) if prices else 0,
            max_purchase_price=max(prices) if prices else 0,
            months_active=len(months),
            first_month=months[0] if months else "",
            last_month=months[-1] if months else "",
            is_entity=_is_entity(name),
            properties=[r.property_address for r in recs],
        ))
    return rankings


# ── Builder / bulk-acquirer detection ─────────────────────────────────
#
# Subdivision developers (e.g. ALSTON TRACE PROPERTIES) appear in the
# Sold Properties feed because they buy raw lots from sellers — but they
# aren't wholesale buyers. We need to keep them out of the main scorecard
# so they don't bury real flippers/landlords behind tract-acquisition
# noise.
#
# Detection uses data signals (not name keywords — too many false
# positives like JCC LEGACY HOMES which is a real flipper):
#   - Same street name appears in 5+ of their deals  → subdivision buy
#   - Same exact sale price appears in 5+ deals      → bulk recording
#   - 10+ deals concentrated in 1 month              → portfolio acquisition
_BULK_SAME_STREET_THRESHOLD = 5
_BULK_SAME_PRICE_THRESHOLD = 5
_BULK_SINGLE_MONTH_DEAL_THRESHOLD = 10


def _street_name(property_address: str) -> str:
    """'5725 Maldon Dr, Louisville, KY' → 'Maldon Dr'.

    Returns the address as-is if no leading house number is present.
    """
    addr = (property_address or "").split(",", 1)[0].strip()
    parts = addr.split(" ", 1)
    if len(parts) > 1 and parts[0].replace("-", "").replace("/", "").isdigit():
        return parts[1].strip()
    return addr


def _classify_buyers(
    rankings: list[BuyerRanking], sold_rows: list[SoldProperty],
) -> None:
    """Classify each buyer as flipper / individual / builder_bulk / excluded.

    Exclusion (state agencies / non-profits) is checked first because such
    entities can also trip the bulk heuristics (Habitat for Humanity buys
    multiple properties on the same street regularly), and we want them
    labeled as "excluded" for transparency rather than "builder_bulk".
    """
    by_buyer: dict[str, list[SoldProperty]] = defaultdict(list)
    for sp in sold_rows:
        by_buyer[(sp.buyer_name or "").strip().upper()].append(sp)

    for r in rankings:
        if _is_excluded(r.buyer_name):
            r.category = "excluded"
            continue
        if not r.is_entity:
            r.category = "individual"
            continue

        recs = by_buyer.get(r.buyer_name, [])
        if not recs:
            r.category = "flipper"
            continue

        street_counts = Counter(_street_name(rec.property_address) for rec in recs)
        price_counts = Counter(rec.sale_amount for rec in recs if rec.sale_amount > 0)
        max_street = max(street_counts.values()) if street_counts else 0
        max_street_name = street_counts.most_common(1)[0][0] if street_counts else ""
        max_price = max(price_counts.values()) if price_counts else 0
        max_price_amount = price_counts.most_common(1)[0][0] if price_counts else 0

        signals = []
        if max_street >= _BULK_SAME_STREET_THRESHOLD:
            signals.append(f"{max_street}x on {max_street_name}")
        if max_price >= _BULK_SAME_PRICE_THRESHOLD:
            signals.append(f"{max_price}x at ${max_price_amount:,}")
        if (r.transaction_count >= _BULK_SINGLE_MONTH_DEAL_THRESHOLD
                and r.months_active <= 1):
            signals.append(f"{r.transaction_count} deals in 1 month")

        if signals:
            r.category = "builder_bulk"
            r.bulk_signal = "; ".join(signals)
        else:
            r.category = "flipper"


# ── Scoring ───────────────────────────────────────────────────────────
# Weights tuned for "actual cash buyer who closes deals at scale":
#   - Frequency (transaction count): biggest signal — pros do volume
#   - Months active (recency / consistency): not a one-off, still buying
#   - Total invested: shows capital deployment
#   - Entity bonus: pros operate through entities
WEIGHT_FREQUENCY = 0.50
WEIGHT_MONTHS_ACTIVE = 0.20
WEIGHT_TOTAL_INVESTED = 0.20
WEIGHT_ENTITY = 0.10


def _score(rankings: list[BuyerRanking]) -> list[BuyerRanking]:
    """Score and rank buyers. Each category gets its own rank space so
    builders and excluded entities don't crowd out wholesale-target buyers
    on the main scorecard.
    """
    if not rankings:
        return rankings

    main = [r for r in rankings if r.category not in ("builder_bulk", "excluded")]
    builders = [r for r in rankings if r.category == "builder_bulk"]
    excluded = [r for r in rankings if r.category == "excluded"]

    for group in (main, builders, excluded):
        if not group:
            continue
        max_count = max(r.transaction_count for r in group) or 1
        max_months = max(r.months_active for r in group) or 1
        max_invested = max(r.total_invested for r in group) or 1
        for r in group:
            freq = (r.transaction_count / max_count) * 100
            months = (r.months_active / max_months) * 100
            invested = (r.total_invested / max_invested) * 100
            entity = 100 if r.is_entity else 0
            r.score = round(
                freq * WEIGHT_FREQUENCY
                + months * WEIGHT_MONTHS_ACTIVE
                + invested * WEIGHT_TOTAL_INVESTED
                + entity * WEIGHT_ENTITY,
                1,
            )
        group.sort(key=lambda r: (-r.score, -r.transaction_count, -r.total_invested))
        for i, r in enumerate(group, 1):
            r.rank = i

    # Combined list — main first (the wholesale targets), then builders,
    # then excluded entities. CLI/Excel filter from this combined list.
    return main + builders + excluded


# ── Excel report ──────────────────────────────────────────────────────


_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
_MONEY_FMT = "#,##0"
_ENTITY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def _write_headers(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_widths(ws, min_w: int = 12, max_w: int = 40) -> None:
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(mx + 2, min_w), max_w)


def generate_buyer_excel(
    rankings: list[BuyerRanking],
    sold_rows: list[SoldProperty],
    *,
    start_month: str,
    end_month: str,
    output_path: str | Path,
    cross_ref: CrossRefResult | None = None,
) -> str:
    """Build the buyer-prospecting workbook and save it.

    If `cross_ref` is provided (from Phase 1B JCD deed cross-reference),
    the main scorecard gains a "Deeds Found" verification column and the
    workbook gets an extra "Deed-Only Buyers" tab listing entity buyers
    in JCD that DataSift's AI missed entirely.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    main_buyers = [r for r in rankings if r.category not in ("builder_bulk", "excluded")]
    builder_buyers = [r for r in rankings if r.category == "builder_bulk"]
    excluded_buyers = [r for r in rankings if r.category == "excluded"]

    # ── Tab 1: Buyer Scorecard (main — wholesale targets only) ────────
    ws = wb.active
    ws.title = "Buyer Scorecard"
    ws.cell(row=1, column=1, value="Jefferson County KY — Cash Buyer Scorecard").font = _TITLE_FONT
    ws.cell(row=2, column=1,
            value=f"Window: {start_month} -> {end_month} · Source: DataSift Sold Properties (Investor tab)"
            ).font = _SUBTITLE_FONT
    ws.cell(row=3, column=1,
            value=(f"{len(main_buyers)} wholesale-target buyers from {len(sold_rows)} sales "
                   f"({len(builder_buyers)} builders/bulk + {len(excluded_buyers)} excluded "
                   f"non-targets separated to their own tabs)")
            ).font = _LABEL_FONT

    # Columns reflect deed verification when cross_ref is provided
    has_deeds = cross_ref is not None
    headers = [
        "Rank", "Buyer Name", "Type", "Score", "# Deals", "Months Active",
        "First Month", "Last Month", "Total Invested", "Avg Price",
        "Min Price", "Max Price",
    ]
    if has_deeds:
        headers.extend(["Deeds Found", "Deed Verified?"])
    _write_headers(ws, 5, headers)
    for i, r in enumerate(main_buyers, 6):
        ws.cell(row=i, column=1, value=r.rank)
        ws.cell(row=i, column=2, value=r.buyer_name)
        type_cell = ws.cell(row=i, column=3, value="Entity" if r.is_entity else "Individual")
        if r.is_entity:
            type_cell.fill = _ENTITY_FILL
        ws.cell(row=i, column=4, value=r.score)
        ws.cell(row=i, column=5, value=r.transaction_count)
        ws.cell(row=i, column=6, value=r.months_active)
        ws.cell(row=i, column=7, value=r.first_month)
        ws.cell(row=i, column=8, value=r.last_month)
        ws.cell(row=i, column=9, value=r.total_invested).number_format = _MONEY_FMT
        ws.cell(row=i, column=10, value=r.avg_purchase_price).number_format = _MONEY_FMT
        ws.cell(row=i, column=11, value=r.min_purchase_price).number_format = _MONEY_FMT
        ws.cell(row=i, column=12, value=r.max_purchase_price).number_format = _MONEY_FMT
        if has_deeds:
            ws.cell(row=i, column=13, value=r.deed_verified_count)
            verified_cell = ws.cell(row=i, column=14, value="Yes" if r.deed_verified_count else "No")
            if r.deed_verified_count:
                verified_cell.fill = _ENTITY_FILL  # reuse green tint for verified
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).border = _THIN_BORDER
    _auto_widths(ws)

    # ── Tab 2: Top Buyer Portfolios (main scorecard top 25) ───────────
    ws2 = wb.create_sheet("Top Buyer Portfolios")
    ws2.cell(row=1, column=1, value="Top 25 Wholesale-Target Buyers — Property-Level Detail").font = _TITLE_FONT
    _write_headers(ws2, 3, ["Rank", "Buyer", "Sale Month", "Property Address", "Sale Amount"])
    row_i = 4
    top25 = {r.buyer_name for r in main_buyers[:25]}
    by_name: dict[str, list[SoldProperty]] = defaultdict(list)
    for sp in sold_rows:
        n = (sp.buyer_name or "").strip().upper()
        if n in top25:
            by_name[n].append(sp)
    for r in main_buyers[:25]:
        for sp in sorted(by_name.get(r.buyer_name, []), key=lambda x: (x.sale_month, x.property_address)):
            ws2.cell(row=row_i, column=1, value=r.rank)
            ws2.cell(row=row_i, column=2, value=sp.buyer_name)
            ws2.cell(row=row_i, column=3, value=sp.sale_month)
            ws2.cell(row=row_i, column=4, value=sp.property_address)
            ws2.cell(row=row_i, column=5, value=sp.sale_amount).number_format = _MONEY_FMT
            for c in range(1, 6):
                ws2.cell(row=row_i, column=c).border = _THIN_BORDER
            row_i += 1
    _auto_widths(ws2)

    # ── Tab 2b: Builders / Bulk Acquirers (separate ranking) ──────────
    ws_b = wb.create_sheet("Builders & Bulk")
    ws_b.cell(row=1, column=1,
              value="Builders / Bulk Acquirers — Separated From Main Scorecard").font = _TITLE_FONT
    ws_b.cell(row=2, column=1,
              value=("Subdivision developers + portfolio bulk buyers. NOT wholesale "
                     "targets. Detected by data signals: 5+ deals on same street, "
                     "5+ deals at identical sale price, or 10+ deals in one month.")
              ).font = _LABEL_FONT
    _write_headers(ws_b, 4, [
        "Rank", "Buyer Name", "Score", "# Deals", "Months Active",
        "First Month", "Last Month", "Total Invested", "Avg Price",
        "Bulk Signal",
    ])
    for i, r in enumerate(builder_buyers, 5):
        ws_b.cell(row=i, column=1, value=r.rank)
        ws_b.cell(row=i, column=2, value=r.buyer_name)
        ws_b.cell(row=i, column=3, value=r.score)
        ws_b.cell(row=i, column=4, value=r.transaction_count)
        ws_b.cell(row=i, column=5, value=r.months_active)
        ws_b.cell(row=i, column=6, value=r.first_month)
        ws_b.cell(row=i, column=7, value=r.last_month)
        ws_b.cell(row=i, column=8, value=r.total_invested).number_format = _MONEY_FMT
        ws_b.cell(row=i, column=9, value=r.avg_purchase_price).number_format = _MONEY_FMT
        ws_b.cell(row=i, column=10, value=r.bulk_signal)
        for c in range(1, 11):
            ws_b.cell(row=i, column=c).border = _THIN_BORDER
    _auto_widths(ws_b)

    # ── Tab 2c: Excluded Entities (state agencies / non-profits) ──────
    ws_x = wb.create_sheet("Excluded Entities")
    ws_x.cell(row=1, column=1,
              value="Excluded Entities — Not Wholesale Targets").font = _TITLE_FONT
    ws_x.cell(row=2, column=1,
              value=("State housing agencies and non-profits that buy properties for "
                     "affordable housing / rehab grants. Listed here for transparency; "
                     "excluded from main scorecard and DataSift CSV. Edit "
                     "_EXCLUDED_BUYER_PATTERNS in jefferson_buyer_prospector.py to add more.")
              ).font = _LABEL_FONT
    _write_headers(ws_x, 4, [
        "Buyer Name", "# Deals", "Months Active", "First Month", "Last Month",
        "Total Invested", "Avg Price",
    ])
    for i, r in enumerate(excluded_buyers, 5):
        ws_x.cell(row=i, column=1, value=r.buyer_name)
        ws_x.cell(row=i, column=2, value=r.transaction_count)
        ws_x.cell(row=i, column=3, value=r.months_active)
        ws_x.cell(row=i, column=4, value=r.first_month)
        ws_x.cell(row=i, column=5, value=r.last_month)
        ws_x.cell(row=i, column=6, value=r.total_invested).number_format = _MONEY_FMT
        ws_x.cell(row=i, column=7, value=r.avg_purchase_price).number_format = _MONEY_FMT
        for c in range(1, 8):
            ws_x.cell(row=i, column=c).border = _THIN_BORDER
    _auto_widths(ws_x)

    # ── Tab 2d: Deed-Only Buyers (DataSift missed) ────────────────────
    # Entity grantees that appear in JCD deed records >= 2 times but are
    # absent from DataSift's Investor list. These are real investors the
    # AI flag missed — the whole point of Phase 1B.
    if cross_ref is not None:
        ws_d = wb.create_sheet("Deed-Only Buyers")
        ws_d.cell(row=1, column=1,
                  value="Deed-Only Buyers — Investors DataSift Missed").font = _TITLE_FONT
        ws_d.cell(row=2, column=1,
                  value=(f"{cross_ref.deed_only_count} wholesale-target entity grantees "
                         f"in Jefferson County deed records (>= 2 filings) that are NOT "
                         f"in DataSift's Investor list. {cross_ref.deed_only_bulk_count} "
                         f"additional bulk acquirers separated to the next tab.")
                  ).font = _LABEL_FONT
        ws_d.cell(row=3, column=1,
                  value=(f"Source: {cross_ref.total_deeds:,} total deeds, "
                         f"{cross_ref.total_entity_grantees:,} entity grantees. "
                         f"DataSift main scorecard: {cross_ref.verified_count} verified, "
                         f"{cross_ref.unverified_count} NOT found in deed records.")
                  ).font = _LABEL_FONT
        _write_headers(ws_d, 5, [
            "Rank", "Buyer Name", "Score", "# Deed Filings", "Months Active",
            "First Filing", "Last Filing", "Sample Legal Description",
        ])
        for i, b in enumerate(cross_ref.deed_only_buyers, 6):
            ws_d.cell(row=i, column=1, value=b.rank)
            ws_d.cell(row=i, column=2, value=b.buyer_name)
            ws_d.cell(row=i, column=3, value=b.score)
            ws_d.cell(row=i, column=4, value=b.deed_count)
            ws_d.cell(row=i, column=5, value=b.months_active)
            ws_d.cell(row=i, column=6, value=b.first_filing)
            ws_d.cell(row=i, column=7, value=b.last_filing)
            ws_d.cell(row=i, column=8,
                      value=b.properties[0] if b.properties else "")
            for c in range(1, 9):
                ws_d.cell(row=i, column=c).border = _THIN_BORDER
        _auto_widths(ws_d)

        # Sibling tab: deed-only BULK acquirers (separated for transparency)
        if cross_ref.deed_only_bulk_buyers:
            ws_db = wb.create_sheet("Deed-Only Bulk")
            ws_db.cell(row=1, column=1,
                       value="Deed-Only Bulk Acquirers — Builders / Estate Sweeps / Portfolios"
                       ).font = _TITLE_FONT
            ws_db.cell(row=2, column=1,
                       value=("Entity grantees with bulk patterns: 5+ deeds with same legal "
                              "description, 5+ deeds filed same day, or 20+ deeds at "
                              "6+/month sustained tempo. Not wholesale targets.")
                       ).font = _LABEL_FONT
            _write_headers(ws_db, 4, [
                "Rank", "Buyer Name", "Score", "# Deed Filings", "Months Active",
                "First Filing", "Last Filing", "Bulk Signal",
            ])
            for i, b in enumerate(cross_ref.deed_only_bulk_buyers, 5):
                ws_db.cell(row=i, column=1, value=b.rank)
                ws_db.cell(row=i, column=2, value=b.buyer_name)
                ws_db.cell(row=i, column=3, value=b.score)
                ws_db.cell(row=i, column=4, value=b.deed_count)
                ws_db.cell(row=i, column=5, value=b.months_active)
                ws_db.cell(row=i, column=6, value=b.first_filing)
                ws_db.cell(row=i, column=7, value=b.last_filing)
                ws_db.cell(row=i, column=8, value=b.bulk_signal)
                for c in range(1, 9):
                    ws_db.cell(row=i, column=c).border = _THIN_BORDER
            _auto_widths(ws_db)

    # ── Tab 3: All Sold Properties (raw) ──────────────────────────────
    ws3 = wb.create_sheet("All Sold (raw)")
    ws3.cell(row=1, column=1, value="Raw Investor Sales — All Months").font = _TITLE_FONT
    _write_headers(ws3, 3, [
        "Sale Month", "Buyer", "Property Address", "County", "Sale Amount",
        "In My Records",
    ])
    for i, sp in enumerate(sorted(sold_rows, key=lambda x: (x.sale_month, x.buyer_name)), 4):
        ws3.cell(row=i, column=1, value=sp.sale_month)
        ws3.cell(row=i, column=2, value=sp.buyer_name)
        ws3.cell(row=i, column=3, value=sp.property_address)
        ws3.cell(row=i, column=4, value=sp.county)
        ws3.cell(row=i, column=5, value=sp.sale_amount).number_format = _MONEY_FMT
        ws3.cell(row=i, column=6, value="Yes" if sp.in_my_records else "")
    _auto_widths(ws3)

    # ── Tab 4: Buyer-Type Mix ─────────────────────────────────────────
    ws4 = wb.create_sheet("Buyer Type Mix")
    ws4.cell(row=1, column=1, value="Entity vs Individual Mix").font = _TITLE_FONT
    entities = sum(1 for r in rankings if r.is_entity)
    individuals = len(rankings) - entities
    entity_invest = sum(r.total_invested for r in rankings if r.is_entity)
    indiv_invest = sum(r.total_invested for r in rankings if not r.is_entity)
    _write_headers(ws4, 3, ["Type", "# Buyers", "Total Invested", "Share of Buyers"])
    total = len(rankings) or 1
    ws4.cell(row=4, column=1, value="Entity (LLC/Trust/Corp)").fill = _ENTITY_FILL
    ws4.cell(row=4, column=2, value=entities)
    ws4.cell(row=4, column=3, value=entity_invest).number_format = _MONEY_FMT
    ws4.cell(row=4, column=4, value=f"{entities/total*100:.1f}%")
    ws4.cell(row=5, column=1, value="Individual")
    ws4.cell(row=5, column=2, value=individuals)
    ws4.cell(row=5, column=3, value=indiv_invest).number_format = _MONEY_FMT
    ws4.cell(row=5, column=4, value=f"{individuals/total*100:.1f}%")
    _auto_widths(ws4)

    wb.save(output_path)
    logger.info("Buyer Excel saved to %s", output_path)
    return str(output_path.resolve())


# ── DataSift-ready CSV ────────────────────────────────────────────────


def export_buyers_csv(rankings: list[BuyerRanking], output_path: str | Path) -> str:
    """Write a DataSift-import CSV for the top 100 buyers.

    No mailing address yet (the Sold Properties page doesn't expose buyer
    mailing addresses). Skip-trace + entity research pipelines fill those
    in downstream.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Builders/bulk acquirers AND state agencies / non-profits are NOT
    # wholesale targets — exclude both categories from the DataSift import.
    target_buyers = [
        r for r in rankings
        if r.category not in ("builder_bulk", "excluded")
    ][:100]
    headers = [
        "owner_name", "address", "city", "state", "zip", "tags", "lists", "notes",
    ]
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for r in target_buyers:
            tags = [
                "cash_buyer", "jefferson_ky",
                f"buyer_score_{int(r.score)}",
                f"deals_{r.transaction_count}",
                "entity" if r.is_entity else "individual",
            ]
            writer.writerow({
                "owner_name": r.buyer_name,
                "address": "",
                "city": "Louisville",
                "state": "KY",
                "zip": "",
                "tags": ",".join(tags),
                "lists": "Cash Buyers - Jefferson",
                "notes": (
                    f"Type: {'Entity' if r.is_entity else 'Individual'}, "
                    f"{r.transaction_count} deals across {r.months_active} months "
                    f"({r.first_month} to {r.last_month}), "
                    f"Total invested ${r.total_invested:,}, "
                    f"Avg ${r.avg_purchase_price:,}"
                ),
            })
    logger.info("Buyer DataSift CSV saved to %s (%d buyers)",
                output_path, len(target_buyers))
    return str(output_path.resolve())


# ── Main entry point ──────────────────────────────────────────────────


def _months_back_to_window(months_back: int) -> tuple[str, str]:
    """Return ('YYYY-MM', 'YYYY-MM') ending with the previous full month.

    We end on the previous full month rather than the current one because
    the current month is partial and would skew per-month aggregates.
    """
    today = datetime.now()
    # Step back to first of current month, then back one more to last full month
    if today.month == 1:
        end_year, end_month = today.year - 1, 12
    else:
        end_year, end_month = today.year, today.month - 1
    end_dt = datetime(end_year, end_month, 1)
    # Step back months_back-1 more months from the end month to get the start
    months_to_step = max(0, months_back - 1)
    sy, sm = end_year, end_month
    for _ in range(months_to_step):
        if sm == 1:
            sy, sm = sy - 1, 12
        else:
            sm -= 1
    return f"{sy:04d}-{sm:02d}", f"{end_year:04d}-{end_month:02d}"


async def run_jefferson_buyer_prospecting(
    *,
    start_month: str | None = None,
    end_month: str | None = None,
    months_back: int = 12,
    output_dir: str | Path | None = None,
    include_deeds: bool = False,
) -> dict:
    """Top-level orchestrator: scrape -> aggregate -> score -> export.

    With include_deeds=True, also scrapes Jefferson County Clerk deed
    transfers for the same window and cross-references the DataSift
    buyers against actual recorded deeds. Adds a deed-verification column
    to the main scorecard and a "Deed-Only Buyers" tab for investors
    DataSift's AI missed entirely.
    """
    if start_month is None or end_month is None:
        start_month, end_month = _months_back_to_window(months_back)
        logger.info(
            "No --start/--end provided, using last %d full months: %s -> %s",
            months_back, start_month, end_month,
        )

    out_dir = Path(output_dir) if output_dir else config.OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    # 1. Scrape
    logger.info("Scraping DataSift Sold Properties: KY Jefferson %s -> %s", start_month, end_month)
    sold_rows = await scrape_sold_properties(
        state_abbr="KY", county_name="Jefferson",
        start_month=start_month, end_month=end_month,
        headless=True,
    )
    if not sold_rows:
        return {"error": "No sold properties scraped — check DataSift login + credentials"}

    # 2. Aggregate, classify, score. Classification (flipper / individual /
    # builder_bulk) MUST run before _score so builders rank in their own
    # ranking space and don't crowd out wholesale-target buyers.
    rankings = aggregate_buyers(sold_rows)
    _classify_buyers(rankings, sold_rows)
    rankings = _score(rankings)

    # 2b. Optional: JCD deed cross-reference. Pulls every recorded deed
    # for the same window, verifies each DataSift buyer appears as a
    # grantee, and surfaces entity buyers DataSift missed entirely.
    deed_transfers: list[DeedTransfer] = []
    cross_ref: CrossRefResult | None = None
    if include_deeds:
        # Convert "YYYY-MM" -> "YYYY-MM-01" for the deed scrape's day-level API
        from calendar import monthrange as _monthrange
        ds_y, ds_m = map(int, end_month.split("-"))
        deed_start = f"{start_month}-01"
        deed_end = f"{end_month}-{_monthrange(ds_y, ds_m)[1]:02d}"
        logger.info("Scraping JCD deed transfers: %s -> %s", deed_start, deed_end)
        # scrape_jefferson_deed_transfers is synchronous (HTTP, not Playwright)
        # — run it in a thread so we don't block the asyncio loop.
        deed_transfers = await asyncio.to_thread(
            scrape_jefferson_deed_transfers, deed_start, deed_end,
        )
        # Cache the deed scrape to disk immediately so a downstream
        # tweak (Excel layout, classifier threshold) doesn't force a
        # re-scrape. ts is set just below; precompute here.
        _ts_for_deeds = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_deed_transfers_csv(
            deed_transfers,
            (Path(output_dir) if output_dir else config.OUTPUT_DIR)
            / f"deeds_KY_Jefferson_{deed_start}_to_{deed_end}_{_ts_for_deeds}.csv",
        )
        cross_ref = cross_reference(rankings, deed_transfers)
        logger.info(
            "Cross-ref: %d verified / %d unverified DataSift buyers; "
            "%d deed-only wholesale targets + %d bulk acquirers DataSift missed",
            cross_ref.verified_count, cross_ref.unverified_count,
            cross_ref.deed_only_count, cross_ref.deed_only_bulk_count,
        )

    # 3. Export
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    raw_csv = export_sold_csv(
        sold_rows,
        out_dir / f"sold_KY_Jefferson_{start_month}_to_{end_month}_{ts}.csv",
    )
    excel_path = generate_buyer_excel(
        rankings, sold_rows,
        start_month=start_month, end_month=end_month,
        output_path=out_dir / f"buyers_KY_Jefferson_{start_month}_to_{end_month}_{ts}.xlsx",
        cross_ref=cross_ref,
    )
    datasift_csv = export_buyers_csv(
        rankings,
        out_dir / f"buyers_datasift_KY_Jefferson_{ts}.csv",
    )

    main_buyers = [r for r in rankings if r.category not in ("builder_bulk", "excluded")]
    builder_buyers = [r for r in rankings if r.category == "builder_bulk"]
    excluded_buyers = [r for r in rankings if r.category == "excluded"]
    by_type = Counter(r.category for r in rankings)
    logger.info(
        "Jefferson buyer prospecting complete: %d total buyers (%d wholesale targets, "
        "%d builders/bulk, %d excluded). Categories: %s",
        len(rankings), len(main_buyers), len(builder_buyers), len(excluded_buyers),
        ", ".join(f"{t}: {c}" for t, c in by_type.items()),
    )

    return {
        "raw_csv": raw_csv,
        "excel": excel_path,
        "datasift_csv": datasift_csv,
        "rankings": rankings,           # main + builders + excluded, in that order
        "main_buyers": main_buyers,
        "builder_buyers": builder_buyers,
        "excluded_buyers": excluded_buyers,
        "sold_rows": sold_rows,
        "deed_transfers": deed_transfers,
        "cross_ref": cross_ref,
        "start_month": start_month,
        "end_month": end_month,
    }

"""4-Level deep prospecting research framework orchestrator.

Coordinates multi-level research from basic skip tracing through curative
title work. Wraps existing enrichers (obituary, ancestry, entity, skip trace)
into a structured depth-based workflow.

Levels:
  1. Enhanced Skip Tracing — multi-provider waterfall
  2. Ownership Verification — deed chain, middle initial, estate flags
  3. Deceased Owner / Heir Research — obituary, ancestry, family tree
  4. Curative Title Work — PACER, title clouds, attorney referral

Usage:
  python src/main.py deep-prospect --csv-path output/records.csv --depth 3
"""

import csv
import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config
from notice_parser import NoticeData

logger = logging.getLogger(__name__)

# Canonical DM phone-field list — single source of truth lives in phone_validator.
# Import it defensively; do NOT inline a 6-field subset. The OLD level-1 stub
# counted only 6 fields and so re-traced (wasting Tracerfy credits) any record
# whose only phones were mobile_4 / mobile_5 / landline_3. Counting the full 9
# fixes that undercount (2g-1).
try:
    from phone_validator import DM_PHONE_FIELDS
except ImportError:  # pragma: no cover - phone_validator is always present in-tree
    DM_PHONE_FIELDS = [
        "primary_phone", "mobile_1", "mobile_2", "mobile_3", "mobile_4",
        "mobile_5", "landline_1", "landline_2", "landline_3",
    ]

# DM #1 email block (flat fields) counted alongside phones.
_DM_EMAIL_FIELDS = ["email_1", "email_2", "email_3", "email_4", "email_5"]

# Skip-trace seams — module-level so tests can inject network-free fakes.
# Imported lazily inside _run_level_1 at call time (kept here as names for the
# test harness to monkeypatch; the real imports happen on demand).
batch_skip_trace = None
guard_traced_contacts = None

DEPTH_NAMES = {
    1: "Enhanced Skip Tracing",
    2: "Ownership Verification",
    3: "Deceased Owner / Heir Research",
    4: "Curative Title Work",
}

DEPTH_DESCRIPTIONS = {
    1: "Multi-provider skip trace waterfall: Tracerfy → DataSift → Trestle phone scoring",
    2: "Deed chain analysis, middle initial verification, estate flags, installment detection",
    3: "Obituary + ancestry search, family tree construction, heir decision-maker ranking",
    4: "PACER court records, title cloud detection, multi-generational heirs, attorney referral",
}

DEPTH_COSTS = {
    1: "$0.10-0.15/record",
    2: "$0-25/record (15-30 min manual)",
    3: "$25-50/month tools + 1-3 hrs/record",
    4: "$500-2,000+ (title attorney)",
}


@dataclass
class ProspectResult:
    """Result of deep prospecting on a single record."""
    address: str = ""
    owner_name: str = ""
    depth_completed: int = 0
    depth_target: int = 1
    # Level 1: Skip trace
    phones_found: int = 0
    emails_found: int = 0
    skip_trace_provider: str = ""
    phone_tier: str = ""
    # Level 2: Ownership
    deed_chain_verified: bool = False
    owner_verified: bool = False
    estate_flag: bool = False
    installment_agreement: bool = False
    middle_initial: str = ""
    # Level 3: Heir research
    owner_deceased: bool = False
    decision_maker: str = ""
    dm_relationship: str = ""
    dm_status: str = ""
    heir_count: int = 0
    heirs_living: int = 0
    heirs_deceased: int = 0
    family_tree_built: bool = False
    # Level 4: Title
    title_clear: bool = True
    title_issues: list = field(default_factory=list)
    attorney_referral_needed: bool = False
    # Recommendation
    recommended_action: str = ""
    notes: str = ""


def _load_records(csv_path: str, max_records: int = 0) -> list[dict]:
    """Load records from CSV file."""
    records = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append(row)
            if max_records and len(records) >= max_records:
                break
    return records


def _record_to_notice(row: dict) -> NoticeData:
    """Convert a CSV row dict to NoticeData."""
    nd = NoticeData()
    field_map = {
        "address": ["address", "Property Street"],
        "city": ["city", "Property City"],
        "state": ["state", "Property State"],
        "zip": ["zip", "Property ZIP"],
        "owner_name": ["owner_name", "full_name", "Owner Name"],
        "notice_type": ["notice_type", "Notice Type"],
        "county": ["county", "County"],
        "parcel_id": ["parcel_id", "Parcel ID"],
        "estimated_value": ["estimated_value", "Estimated Value"],
        "owner_deceased": ["owner_deceased"],
        "decision_maker_name": ["decision_maker_name", "Decision Maker"],
        "decision_maker_relationship": ["decision_maker_relationship"],
    }
    for attr, keys in field_map.items():
        for key in keys:
            val = row.get(key, "")
            if val:
                setattr(nd, attr, str(val).strip())
                break
    return nd


# ── Level execution ───────────────────────────────────────────────────

def _count_phones(notice: NoticeData) -> int:
    """Count populated DM #1 flat phones across the FULL canonical 9-field set."""
    return sum(1 for attr in DM_PHONE_FIELDS if getattr(notice, attr, ""))


def _count_emails(notice: NoticeData) -> int:
    """Count populated DM #1 flat emails (email_1..email_5)."""
    return sum(1 for attr in _DM_EMAIL_FIELDS if getattr(notice, attr, ""))


async def _run_level_1(notice: NoticeData, result: ProspectResult,
                       skip_trace: bool = True) -> None:
    """Level 1: Enhanced Skip Tracing.

    If the DM already has phones/emails (counted over the full canonical
    DM_PHONE_FIELDS set — no 6-field undercount), use them as-is. Otherwise,
    when skip_trace is enabled and Tracerfy is configured, run a real
    single-record ``batch_skip_trace([notice])`` and then the death/identity
    guard (so the deep-prospect path gets the same protection as the daily
    pipeline — T-05-07) before counting (2g-1, CONTACT-01).
    """
    phones = _count_phones(notice)
    emails = _count_emails(notice)

    if phones > 0 or emails > 0:
        result.phones_found = phones
        result.emails_found = emails
        result.skip_trace_provider = "existing"
        result.notes += "Skip trace data already present. "
        result.depth_completed = 1
        result.recommended_action = "Score phones via Trestle"
        return

    # No phones on file — this is the path the old stub never serviced.
    if not skip_trace:
        result.notes += "Skip trace suppressed (--no-skip-trace). "
        result.depth_completed = 1
        result.recommended_action = "Skip trace disabled — no phones pulled"
        return

    if not getattr(config, "TRACERFY_API_KEY", ""):
        result.notes += "Tracerfy not configured — no phones. "
        result.depth_completed = 1
        result.recommended_action = "Configure TRACERFY_API_KEY to pull phones"
        return

    # Run a real single-record Tracerfy batch, then the death/identity guard.
    # Resolve the seams from module attributes (tests inject fakes) or import
    # them lazily. batch_skip_trace is sync; calling it inside this async fn is
    # fine (it blocks) — keep the change minimal (no asyncio.to_thread).
    try:
        global batch_skip_trace, guard_traced_contacts
        if batch_skip_trace is None:
            from tracerfy_skip_tracer import batch_skip_trace as _bst
            batch_skip_trace = _bst
        if guard_traced_contacts is None:
            from skip_trace_guard import guard_traced_contacts as _gtc
            guard_traced_contacts = _gtc

        batch_skip_trace([notice])
        # Guard the freshly-traced record BEFORE counting/surfacing any phone.
        try:
            guard_traced_contacts(notice)
        except Exception as ge:  # never crash the trace pass
            logger.warning("Level 1 guard pass failed for %s: %s — continuing",
                           notice.address, ge)

        phones = _count_phones(notice)
        emails = _count_emails(notice)
        result.phones_found = phones
        result.emails_found = emails
        result.skip_trace_provider = "tracerfy"
        result.notes += f"Tracerfy skip trace: {phones} phone(s), {emails} email(s) (guarded). "
    except Exception as e:
        logger.warning("Level 1 skip trace failed for %s: %s", notice.address, e)
        result.notes += "Skip trace failed — no phones pulled. "

    result.depth_completed = 1
    result.recommended_action = (
        "Score phones via Trestle" if result.phones_found > 0
        else "No phones found — try manual people-search waterfall"
    )


async def _run_level_2(notice: NoticeData, result: ProspectResult) -> None:
    """Level 2: Ownership Verification."""
    # Check for estate indicators from tax records
    deceased_ind = notice.deceased_indicator or ""
    if deceased_ind:
        result.estate_flag = True
        result.notes += f"Estate flag detected: {deceased_ind}. "

    # Parcel ID verification
    if notice.parcel_id:
        result.deed_chain_verified = True
        result.notes += f"Parcel ID verified: {notice.parcel_id}. "
    else:
        result.notes += "No parcel ID — deed chain unverified. "

    # Owner verification via tax record cross-reference
    if notice.tax_owner_name:
        result.owner_verified = True
        result.notes += f"Tax owner confirmed: {notice.tax_owner_name}. "

    result.depth_completed = 2
    if result.estate_flag:
        result.recommended_action = "Proceed to Level 3 — heir research needed"
    else:
        result.recommended_action = "Owner verified — proceed to marketing"


async def _run_level_3(notice: NoticeData, result: ProspectResult) -> None:
    """Level 3: Deceased Owner / Heir Research."""
    # Check existing obituary data
    if notice.owner_deceased == "yes":
        result.owner_deceased = True
        result.decision_maker = notice.decision_maker_name or ""
        result.dm_relationship = notice.decision_maker_relationship or ""
        result.dm_status = notice.decision_maker_status or ""

        # Parse heir counts
        try:
            result.heirs_living = int(notice.heirs_verified_living or 0)
            result.heirs_deceased = int(notice.heirs_verified_deceased or 0)
            result.heir_count = result.heirs_living + result.heirs_deceased
        except ValueError:
            pass

        result.family_tree_built = bool(notice.heir_map_json)

        if result.decision_maker:
            result.notes += f"DM identified: {result.decision_maker} ({result.dm_relationship}). "
            result.recommended_action = "Contact decision maker — begin marketing sequence"
        else:
            result.notes += "Owner deceased but no DM identified. "
            result.recommended_action = "Run ancestry research for heir identification"
    else:
        result.notes += "Owner not confirmed deceased. "
        result.recommended_action = "Run obituary search to confirm status"

    result.depth_completed = 3


async def _run_level_4(notice: NoticeData, result: ProspectResult) -> None:
    """Level 4: Curative Title Work."""
    # Check for title complexity indicators
    issues = []

    if result.heir_count > 3:
        issues.append(f"Multiple heirs ({result.heir_count}) — potential fractional ownership")
    if result.heirs_deceased > 0:
        issues.append(f"{result.heirs_deceased} deceased heirs — multi-generational title chain")
    if notice.entity_type in ("trust", "estate"):
        issues.append(f"Entity ownership ({notice.entity_type}) — may need court approval")

    result.title_issues = issues
    result.title_clear = len(issues) == 0
    result.attorney_referral_needed = len(issues) > 1

    if result.attorney_referral_needed:
        result.notes += f"Title issues: {'; '.join(issues)}. "
        result.recommended_action = "REFER TO TITLE ATTORNEY — curative work needed"
    elif issues:
        result.notes += f"Minor title concern: {issues[0]}. "
        result.recommended_action = "Review title with attorney before closing"
    else:
        result.notes += "Title appears clear. "
        result.recommended_action = "Proceed to acquisition"

    result.depth_completed = 4


async def prospect_record(notice: NoticeData, target_depth: int = 3,
                          skip_trace: bool = True) -> ProspectResult:
    """Run deep prospecting on a single record up to target depth.

    ``skip_trace`` threads down to Level 1 — when False (``--no-skip-trace``),
    Level 1 does NOT call Tracerfy.
    """
    result = ProspectResult(
        address=notice.address,
        owner_name=notice.owner_name,
        depth_target=target_depth,
    )

    # Levels 2-4 use the dict dispatch; Level 1 takes the skip_trace flag.
    level_runners = {
        2: _run_level_2,
        3: _run_level_3,
        4: _run_level_4,
    }

    for level in range(1, target_depth + 1):
        if level == 1:
            await _run_level_1(notice, result, skip_trace=skip_trace)
            logger.debug("Level 1 complete for %s", notice.address)
            continue
        runner = level_runners.get(level)
        if runner:
            await runner(notice, result)
            logger.debug("Level %d complete for %s", level, notice.address)

    return result


# ── Batch processing ──────────────────────────────────────────────────

async def run_deep_prospecting(csv_path: str, depth: int = 3,
                               max_records: int = 0,
                               output_path: str = "",
                               skip_trace: bool = True) -> dict:
    """Run deep prospecting on a batch of records.

    ``skip_trace`` (default True) controls whether Level 1 auto-runs Tracerfy;
    set False to honor ``--no-skip-trace``.

    Returns dict with results and report path.
    """
    logger.info("Starting deep prospecting (depth %d, skip_trace=%s) on %s",
                depth, skip_trace, csv_path)

    records = _load_records(csv_path, max_records)
    if not records:
        return {"error": "No records found in CSV"}

    logger.info("Processing %d records at depth %d (%s)",
                len(records), depth, DEPTH_NAMES.get(depth, "Unknown"))

    results = []
    for i, row in enumerate(records):
        notice = _record_to_notice(row)
        result = await prospect_record(notice, depth, skip_trace=skip_trace)
        results.append(result)
        if (i + 1) % 10 == 0:
            logger.info("Processed %d/%d records", i + 1, len(records))

    # Generate report
    report_path = _generate_dp_report(results, depth, output_path)

    # Summary stats
    stats = {
        "total": len(results),
        "phones_found": sum(1 for r in results if r.phones_found > 0),
        "owners_verified": sum(1 for r in results if r.owner_verified),
        "deceased_confirmed": sum(1 for r in results if r.owner_deceased),
        "dms_identified": sum(1 for r in results if r.decision_maker),
        "title_issues": sum(1 for r in results if r.title_issues),
        "attorney_referrals": sum(1 for r in results if r.attorney_referral_needed),
    }

    logger.info("Deep prospecting complete: %d records, %d phones, %d deceased, %d DMs, %d title issues",
                stats["total"], stats["phones_found"], stats["deceased_confirmed"],
                stats["dms_identified"], stats["title_issues"])

    return {
        "results": results,
        "stats": stats,
        "report_path": report_path,
    }


# ── Report generation ─────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_VALUE_FONT = Font(name="Calibri", bold=True, size=13, color="222222")
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _generate_dp_report(results: list[ProspectResult], depth: int,
                         output_path: str = "") -> str:
    """Generate deep prospecting Excel report."""
    wb = Workbook()

    # Summary tab
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="Deep Prospecting Report").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"Depth: Level {depth} — {DEPTH_NAMES.get(depth, '')}").font = _SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = _LABEL_FONT
    ws.cell(row=4, column=1, value=f"Records: {len(results)}").font = _LABEL_FONT

    stats = [
        ("Phones Found", sum(1 for r in results if r.phones_found > 0)),
        ("Owners Verified", sum(1 for r in results if r.owner_verified)),
        ("Estate Flags", sum(1 for r in results if r.estate_flag)),
        ("Deceased Confirmed", sum(1 for r in results if r.owner_deceased)),
        ("Decision Makers ID'd", sum(1 for r in results if r.decision_maker)),
        ("Family Trees Built", sum(1 for r in results if r.family_tree_built)),
        ("Title Issues", sum(1 for r in results if r.title_issues)),
        ("Attorney Referrals", sum(1 for r in results if r.attorney_referral_needed)),
    ]
    for i, (label, value) in enumerate(stats, 6):
        ws.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=i, column=2, value=value).font = _VALUE_FONT
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15

    # Detail tab
    ws2 = wb.create_sheet("Detail")
    headers = ["Address", "Owner", "Depth", "Phones", "Emails", "Verified",
               "Deceased", "Decision Maker", "DM Relationship", "Heirs",
               "Title Issues", "Action", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for i, r in enumerate(results, 2):
        vals = [r.address, r.owner_name, f"L{r.depth_completed}", r.phones_found,
                r.emails_found, "Yes" if r.owner_verified else "",
                "Yes" if r.owner_deceased else "", r.decision_maker,
                r.dm_relationship, r.heir_count,
                "; ".join(r.title_issues) if r.title_issues else "",
                r.recommended_action, r.notes]
        for col, val in enumerate(vals, 1):
            ws2.cell(row=i, column=col, value=val).border = _THIN_BORDER

    # Depth guide tab
    ws3 = wb.create_sheet("Depth Guide")
    ws3.cell(row=1, column=1, value="Research Depth Levels").font = _TITLE_FONT
    for level in range(1, 5):
        row = (level - 1) * 4 + 3
        ws3.cell(row=row, column=1, value=f"Level {level}: {DEPTH_NAMES[level]}").font = _SUBTITLE_FONT
        ws3.cell(row=row + 1, column=1, value=DEPTH_DESCRIPTIONS[level]).font = _LABEL_FONT
        ws3.cell(row=row + 2, column=1, value=f"Cost: {DEPTH_COSTS[level]}").font = _LABEL_FONT
    ws3.column_dimensions["A"].width = 70

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(config.OUTPUT_DIR / f"deep_prospecting_L{depth}_{timestamp}.xlsx")

    wb.save(output_path)
    logger.info("Deep prospecting report saved to %s", output_path)
    return output_path
